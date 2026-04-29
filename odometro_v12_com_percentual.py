import os
import sys
import pandas as pd
import numpy as np


# ============================================================
# V12 - ODOMETRO_MATCH UNICO + SELECAO MANUAL + % ANDAMENTO
# ============================================================
# OBJETIVO:
# - Permitir escolher os arquivos base em qualquer pasta.
# - Mostrar qual arquivo deve ser selecionado em cada etapa.
# - Usar AM/PM somente como regra interna de busca, sem gerar coluna AM/PM.
# - Gerar apenas uma coluna oficial de odometro: ODOMETRO_MATCH.
#
# BASES NECESSARIAS:
# 1) Base de dados - Combustivel.xlsx
# 2) Base de dados - Km Rodado Maxtrack.xlsx
# 3) Base de dados - Ativo de Veiculos.xlsx
# 4) Producao Oficial - Relatorio Cliente.xlsx
#
# REGRAS PRINCIPAIS:
# - Maxtrack e a fonte oficial de odometro.
# - Quando nao encontrar Maxtrack, usar distancia do Cliente ate voltar ao proximo Maxtrack.
# - ODOMETRO_MATCH nao pode ter salto maior que 1500 entre abastecimentos.
# - KM entre abastecimentos deve ser igual ou superior a distancia do Cliente no intervalo.
# - Nao gerar ODOMETRO_FINAL, ODOMETRO_CALCULADO ou outra variacao.
# ============================================================


# ============================================================
# CONFIG
# ============================================================
ABA_COMBUSTIVEL = "Abastecimentos"
ABA_MAXTRACK = "RL - Viagens"
ABA_ATIVO = "Ativo atualizado"
ABA_PRODUCAO = "BD Transporte"

# Agrupamento de abastecimentos proximos
MINUTOS_AGRUPAMENTO = 30
MAX_DIFF_ODOM_AGRUPAMENTO = 160

# Busca Maxtrack
JANELA_MATCH_MINUTOS = 360
TOP_CANDIDATOS_POR_ABAST = 10

# Regras duras
MAX_SALTO_ABAST = 1500
MIN_INCREMENTO_ABSOLUTO = 2
MIN_INCREMENTO_SEM_CLIENTE = 20
MAX_VARIACAO_LOCAL_SEM_CLIENTE = 800
JANELA_TENDENCIA = 3

# Tolerancias
TOLERANCIA_ANCORA_ABS = 80
TOLERANCIA_ANCORA_PCT = 0.08
MARGEM_TECNICA_CLIENTE = 0

# Score interno
PESO_MESMO_DIA = 3.0
PESO_MESMO_PERIODO = 3.0
PESO_TEMPO = 2.0
PESO_TIPO_REF = 0.8
PESO_CLIENTE = 3.0
PESO_CORREDOR = 4.0


# ============================================================
# PROGRESSO NO TERMINAL
# ============================================================
def mostrar_progresso(etapa, atual, total):
    """
    Mostra percentual de andamento no terminal.
    Nao interfere no processamento.
    """
    if total <= 0:
        return

    pct = (atual / total) * 100
    barra_total = 30
    preenchido = int(barra_total * atual / total)
    barra = "#" * preenchido + "-" * (barra_total - preenchido)

    print(f"\r{etapa}: [{barra}] {pct:6.2f}% ({atual}/{total})", end="", flush=True)

    if atual >= total:
        print()


def imprimir_etapa(texto):
    print("\n" + "=" * 70)
    print(texto)
    print("=" * 70)



# ============================================================
# SELECAO DE ARQUIVOS
# ============================================================
def selecionar_arquivo_gui(titulo, descricao):
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)

        messagebox.showinfo("Selecionar arquivo", descricao)

        caminho = filedialog.askopenfilename(
            title=titulo,
            filetypes=[
                ("Arquivos Excel", "*.xlsx *.xls"),
                ("Todos os arquivos", "*.*"),
            ],
        )

        root.destroy()

        if not caminho:
            print("Nenhum arquivo selecionado. Execucao cancelada.")
            sys.exit(0)

        return caminho

    except Exception:
        print("\n" + "=" * 70)
        print(descricao)
        print("=" * 70)
        caminho = input("Cole o caminho completo do arquivo Excel: ").strip().strip('"')
        if not caminho:
            print("Nenhum arquivo informado. Execucao cancelada.")
            sys.exit(0)
        return caminho


def selecionar_saida_gui(sugestao):
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)

        messagebox.showinfo(
            "Salvar resultado",
            "Agora escolha onde salvar o arquivo final gerado.\n\n"
            "Sugestao: mantenha o nome Resumo_Abastecimento_Odometro_V12.xlsx"
        )

        caminho = filedialog.asksaveasfilename(
            title="Salvar arquivo final",
            initialfile=sugestao,
            defaultextension=".xlsx",
            filetypes=[
                ("Arquivos Excel", "*.xlsx"),
                ("Todos os arquivos", "*.*"),
            ],
        )

        root.destroy()

        if not caminho:
            print("Nenhum local de saida selecionado. Execucao cancelada.")
            sys.exit(0)

        return caminho

    except Exception:
        print("\nEscolha onde salvar o arquivo final.")
        caminho = input(f"Caminho de saida sugerido [{sugestao}]: ").strip().strip('"')
        return caminho if caminho else os.path.abspath(sugestao)


def selecionar_bases():
    print("\n" + "=" * 70)
    print("SELECAO DAS BASES")
    print("=" * 70)
    print("Voce pode escolher os arquivos em qualquer pasta.")
    print("O sistema vai pedir um arquivo por vez e informar qual base selecionar.")
    print("=" * 70)

    arq_combustivel = selecionar_arquivo_gui(
        "1 de 4 - Selecione a BASE DE COMBUSTIVEL",
        "1 de 4\n\nSelecione a BASE DE COMBUSTIVEL.\n\n"
        "Arquivo esperado normalmente:\n"
        "Base de dados - Combustivel.xlsx\n\n"
        "Aba esperada:\n"
        f"{ABA_COMBUSTIVEL}\n\n"
        "Essa base deve conter os abastecimentos, veiculo, data e litros."
    )

    arq_maxtrack = selecionar_arquivo_gui(
        "2 de 4 - Selecione a BASE MAXTRACK",
        "2 de 4\n\nSelecione a BASE MAXTRACK / KM RODADO.\n\n"
        "Arquivo esperado normalmente:\n"
        "Base de dados - Km Rodado Maxtrack.xlsx\n\n"
        "Aba esperada:\n"
        f"{ABA_MAXTRACK}\n\n"
        "Essa base deve conter placa, inicio, fim, odometro inicial e odometro final."
    )

    arq_ativo = selecionar_arquivo_gui(
        "3 de 4 - Selecione a BASE DE ATIVOS",
        "3 de 4\n\nSelecione a BASE DE ATIVOS / VEICULOS.\n\n"
        "Arquivo esperado normalmente:\n"
        "Base de dados - Ativo de Veiculos.xlsx\n\n"
        "Aba esperada:\n"
        f"{ABA_ATIVO}\n\n"
        "Essa base deve conter GO e PLACA para vincular combustivel com placa."
    )

    arq_producao = selecionar_arquivo_gui(
        "4 de 4 - Selecione a BASE CLIENTE / PRODUCAO",
        "4 de 4\n\nSelecione a BASE CLIENTE / PRODUCAO OFICIAL.\n\n"
        "Arquivo esperado normalmente:\n"
        "Producao Oficial - Relatorio Cliente.xlsx\n\n"
        "Aba esperada:\n"
        f"{ABA_PRODUCAO}\n\n"
        "Essa base deve conter placa, saida da viagem e distancia percorrida."
    )

    pasta_saida = os.path.dirname(arq_combustivel) or os.getcwd()
    sugestao_saida = os.path.join(pasta_saida, "Resumo_Abastecimento_Odometro.xlsx")
    arq_saida = selecionar_saida_gui(sugestao_saida)

    print("\nArquivos selecionados:")
    print(f"Combustivel: {arq_combustivel}")
    print(f"Maxtrack:    {arq_maxtrack}")
    print(f"Ativo:       {arq_ativo}")
    print(f"Cliente:     {arq_producao}")
    print(f"Saida:       {arq_saida}")
    print("=" * 70)

    return arq_combustivel, arq_maxtrack, arq_ativo, arq_producao, arq_saida


# ============================================================
# FUNCOES AUXILIARES
# ============================================================
def normalizar_nome_coluna(valor):
    if valor is None:
        return ""
    txt = str(valor).strip().upper()
    txt = txt.replace("\n", " ").replace("\r", " ").replace("\t", " ")
    txt = " ".join(txt.split())
    txt = txt.replace(" ", "").replace("-", "").replace("_", "")
    return txt


def normalizar_placa(valor):
    if pd.isna(valor):
        return None
    txt = str(valor).strip().upper()
    txt = txt.replace("-", "").replace(" ", "")
    return txt if txt else None


def extrair_go_do_veiculo(valor):
    if pd.isna(valor):
        return None
    txt = "".join(ch for ch in str(valor) if ch.isdigit())
    if not txt:
        return None
    return txt[-5:]


def para_numero(valor):
    if pd.isna(valor):
        return np.nan
    if isinstance(valor, (int, float, np.integer, np.floating)):
        return float(valor)

    txt = str(valor).strip()

    # Trata formato brasileiro: 1.234,56
    try:
        return float(txt.replace(".", "").replace(",", "."))
    except Exception:
        pass

    try:
        return float(txt)
    except Exception:
        return np.nan


def escolher_coluna(df, candidatos, obrigatoria=True):
    cols_original = list(df.columns)
    cols_norm = {normalizar_nome_coluna(c): c for c in cols_original}

    for candidato in candidatos:
        cand_norm = normalizar_nome_coluna(candidato)

        if cand_norm in cols_norm:
            return cols_norm[cand_norm]

        for col_norm, col_real in cols_norm.items():
            if cand_norm in col_norm or col_norm in cand_norm:
                return col_real

    if obrigatoria:
        print("\nColunas disponiveis:")
        for c in cols_original:
            print(f" - {c}")
        raise KeyError(f"Nenhuma das colunas foi encontrada: {candidatos}")
    return None


def escolher_melhor_coluna_por_palavras(df, lista_de_palavras, obrigatoria=True):
    cols_original = list(df.columns)

    def quebrar(txt):
        base = str(txt).upper()
        base = base.replace("\n", " ").replace("\r", " ").replace("\t", " ")
        base = base.replace("-", " ").replace("_", " ").replace("/", " ")
        base = " ".join(base.split())
        return set(base.split())

    palavras_alvo = set(p.upper() for p in lista_de_palavras)
    melhor_coluna = None
    melhor_score = -1

    for col in cols_original:
        palavras_col = quebrar(col)
        score = len(palavras_alvo.intersection(palavras_col))
        if score > melhor_score:
            melhor_score = score
            melhor_coluna = col

    if melhor_score > 0:
        return melhor_coluna

    if obrigatoria:
        print("\nColunas disponiveis:")
        for c in cols_original:
            print(f" - {c}")
        raise KeyError(f"Nenhuma coluna compativel com palavras: {lista_de_palavras}")

    return None


def ler_excel_com_aba(caminho, aba_preferida):
    try:
        return pd.read_excel(caminho, sheet_name=aba_preferida)
    except Exception:
        xls = pd.ExcelFile(caminho)
        abas = xls.sheet_names

        print(f"\nNao foi possivel ler a aba '{aba_preferida}' no arquivo:")
        print(caminho)
        print("\nAbas encontradas:")
        for idx, aba in enumerate(abas, start=1):
            print(f"{idx} - {aba}")

        if len(abas) == 1:
            print(f"\nUsando a unica aba encontrada: {abas[0]}")
            return pd.read_excel(caminho, sheet_name=abas[0])

        escolha = input("\nDigite o numero da aba correta: ").strip()
        try:
            pos = int(escolha) - 1
            aba = abas[pos]
        except Exception:
            raise ValueError("Aba invalida selecionada.")

        return pd.read_excel(caminho, sheet_name=aba)


def diferenca_minutos(dt1, dt2):
    if pd.isna(dt1) or pd.isna(dt2):
        return np.nan
    return abs((dt1 - dt2).total_seconds()) / 60.0


def periodo_interno(dt):
    if pd.isna(dt):
        return None
    return "AM" if dt.hour < 12 else "PM"


def mesmo_periodo_interno(dt1, dt2):
    p1 = periodo_interno(dt1)
    p2 = periodo_interno(dt2)
    return p1 is not None and p1 == p2


def classificar_confianca_temporal(dif_min):
    if pd.isna(dif_min):
        return "NENHUM"
    if dif_min <= 10:
        return "ALTA"
    if dif_min <= 30:
        return "MEDIA"
    if dif_min <= 120:
        return "BAIXA"
    return "MUITO_BAIXA"


def somar_distancia_producao(df_prod_placa, dt_ini, dt_fim):
    if df_prod_placa is None or df_prod_placa.empty:
        return np.nan
    if pd.isna(dt_ini) or pd.isna(dt_fim) or dt_fim < dt_ini:
        return np.nan

    janela = df_prod_placa[
        (df_prod_placa["DATA_HORA_PROD"] > dt_ini) &
        (df_prod_placa["DATA_HORA_PROD"] <= dt_fim)
    ].copy()

    if janela.empty:
        return np.nan

    return float(janela["DISTANCIA_PROD"].sum())


def mediana_km_validos(grupo, idx_atual, janela=JANELA_TENDENCIA):
    if "KM_ENTRE_ABAST" not in grupo.columns:
        return np.nan

    inicio = max(0, idx_atual - janela)
    sub = grupo.loc[inicio:idx_atual - 1].copy()
    vals = sub["KM_ENTRE_ABAST"].dropna().tolist()
    vals = [float(v) for v in vals if float(v) > 1 and float(v) <= MAX_SALTO_ABAST]

    if not vals:
        return np.nan

    return float(np.median(vals))


def classificar_diff_mes(x):
    if pd.isna(x):
        return "SEM_COMPARACAO"
    x = abs(float(x))
    if x <= 100:
        return "OTIMA"
    if x <= 250:
        return "BOA"
    if x <= 500:
        return "ATENCAO"
    return "CRITICA"


# ============================================================
# CANDIDATOS MAXTRACK
# ============================================================
def listar_candidatos_match(df_placa, dt_abast):
    if df_placa is None or df_placa.empty or pd.isna(dt_abast):
        return pd.DataFrame()

    inicio = dt_abast - pd.Timedelta(minutes=JANELA_MATCH_MINUTOS)
    fim = dt_abast + pd.Timedelta(minutes=JANELA_MATCH_MINUTOS)

    janela = df_placa[
        (df_placa["DATA_HORA_REF"] >= inicio) &
        (df_placa["DATA_HORA_REF"] <= fim)
    ].copy()

    if janela.empty:
        return janela

    janela["DIFERENCA_MINUTOS"] = (
        (janela["DATA_HORA_REF"] - dt_abast).abs().dt.total_seconds() / 60.0
    )

    # AM/PM usado apenas internamente. Nao sera exportado.
    janela["_MESMO_DIA_INTERNO"] = janela["DATA_HORA_REF"].dt.date == dt_abast.date()
    janela["_MESMO_PERIODO_INTERNO"] = janela["DATA_HORA_REF"].apply(lambda x: mesmo_periodo_interno(dt_abast, x))

    # Score inicial: tempo + mesmo dia + mesmo periodo + tipo FIM/INICIO
    janela["_SCORE_INICIAL"] = 0.0
    janela["_SCORE_INICIAL"] += (1 - (janela["DIFERENCA_MINUTOS"].clip(0, JANELA_MATCH_MINUTOS) / JANELA_MATCH_MINUTOS)) * PESO_TEMPO
    janela["_SCORE_INICIAL"] += np.where(janela["_MESMO_DIA_INTERNO"], PESO_MESMO_DIA, 0)
    janela["_SCORE_INICIAL"] += np.where(janela["_MESMO_PERIODO_INTERNO"], PESO_MESMO_PERIODO, 0)
    janela["_SCORE_INICIAL"] += np.where(janela["TIPO_REF"] == "FIM", PESO_TIPO_REF, PESO_TIPO_REF * 0.75)

    janela = janela.sort_values(
        by=["_SCORE_INICIAL", "DIFERENCA_MINUTOS", "DATA_HORA_REF"],
        ascending=[False, True, False]
    ).head(TOP_CANDIDATOS_POR_ABAST).reset_index(drop=True)

    return janela


# ============================================================
# CONSOLIDACAO DE ABASTECIMENTOS
# ============================================================
def agrupar_abastecimentos_proximos(grupo):
    grupo = grupo.sort_values("DATA_HORA_ABASTECIMENTO").copy().reset_index(drop=True)
    if grupo.empty:
        return grupo

    grupos = []
    atual = [0]

    for i in range(1, len(grupo)):
        row_ref = grupo.loc[atual[-1]]
        row_cur = grupo.loc[i]

        delta_min = (
            (row_cur["DATA_HORA_ABASTECIMENTO"] - row_ref["DATA_HORA_ABASTECIMENTO"]).total_seconds() / 60.0
        )

        odom_ref = row_ref["ODOMETRO_CANDIDATO"]
        odom_cur = row_cur["ODOMETRO_CANDIDATO"]

        pode_agrupar = False
        if delta_min <= MINUTOS_AGRUPAMENTO:
            if pd.notna(odom_ref) and pd.notna(odom_cur):
                diff_odom = abs(float(odom_cur) - float(odom_ref))
                if diff_odom <= MAX_DIFF_ODOM_AGRUPAMENTO:
                    pode_agrupar = True
            else:
                pode_agrupar = True

        if pode_agrupar:
            atual.append(i)
        else:
            grupos.append(atual)
            atual = [i]

    if atual:
        grupos.append(atual)

    linhas = []

    for gid, idxs in enumerate(grupos, start=1):
        bloco = grupo.loc[idxs].copy()

        odom_validos = bloco["ODOMETRO_CANDIDATO"].dropna()
        dt_mt_validos = bloco["DATA_HORA_MAXTRACK_ENCONTRADA"].dropna()
        dif_validos = bloco["DIFERENCA_MINUTOS"].dropna()

        confs = bloco["CONFIANCA_TEMPORAL"].dropna().tolist()
        if "ALTA" in confs:
            conf_grupo = "ALTA"
        elif "MEDIA" in confs:
            conf_grupo = "MEDIA"
        elif "BAIXA" in confs:
            conf_grupo = "BAIXA"
        elif "MUITO_BAIXA" in confs:
            conf_grupo = "MUITO_BAIXA"
        else:
            conf_grupo = "NENHUM"

        linhas.append({
            "PLACA": bloco["PLACA"].iloc[0],
            "GO": bloco["GO"].iloc[0],
            "VEICULO": bloco["VEICULO"].iloc[0],
            "DATA_HORA_ABASTECIMENTO": bloco["DATA_HORA_ABASTECIMENTO"].min(),
            "LITROS": bloco["LITROS"].sum(),
            "ODOMETRO_MATCH": np.nan,
            "ODOMETRO_CANDIDATO": odom_validos.min() if len(odom_validos) > 0 else np.nan,
            "DATA_HORA_MAXTRACK_ENCONTRADA": dt_mt_validos.iloc[0] if len(dt_mt_validos) > 0 else pd.NaT,
            "TIPO_REF_MAXTRACK": bloco["TIPO_REF_MAXTRACK"].dropna().iloc[0] if bloco["TIPO_REF_MAXTRACK"].notna().any() else None,
            "DIFERENCA_MINUTOS": dif_validos.min() if len(dif_validos) > 0 else np.nan,
            "STATUS_MATCH": "ENCONTRADO" if len(odom_validos) > 0 else bloco["STATUS_MATCH"].iloc[0],
            "CONFIANCA_TEMPORAL": conf_grupo,
            "QTD_EVENTOS_GRUPO": len(bloco),
            "AGRUPADO_DUPLICADO": "SIM" if len(bloco) > 1 else "NAO",
            "GRUPO_ABASTECIMENTO": gid,
            "IDS_LINHAS_ORIGEM": ",".join(str(x) for x in bloco["ID_LINHA_ORIGINAL"].tolist())
        })

    return pd.DataFrame(linhas)


# ============================================================
# DEFINICAO DE ANCORAS
# ============================================================
def localizar_ancora_anterior(grupo, idx):
    sub = grupo.loc[:idx - 1]
    sub = sub[sub["EH_ANCORA"] == True]
    if sub.empty:
        return None
    return int(sub.index.max())


def localizar_ancora_posterior(grupo, idx):
    sub = grupo.loc[idx + 1:]
    sub = sub[sub["EH_ANCORA"] == True]
    if sub.empty:
        return None
    return int(sub.index.min())


def avaliar_candidato_contexto(grupo, idx, cand, df_prod_placa):
    od_cand = float(cand["ODOMETRO"])
    dt_cand = cand["DATA_HORA_REF"]
    dt_abast = grupo.loc[idx, "DATA_HORA_ABASTECIMENTO"]

    score = float(cand.get("_SCORE_INICIAL", 0.0))
    motivos = []

    idx_prev = localizar_ancora_anterior(grupo, idx)
    idx_next = localizar_ancora_posterior(grupo, idx)

    if idx_prev is not None:
        od_prev = grupo.loc[idx_prev, "ODOMETRO_MATCH"]
        dt_prev = grupo.loc[idx_prev, "DATA_HORA_ABASTECIMENTO"]

        if pd.notna(od_prev):
            delta = od_cand - float(od_prev)
            km_cli = somar_distancia_producao(df_prod_placa, dt_prev, dt_abast)

            if delta <= 1:
                motivos.append("NAO_EVOLUIU_DESDE_ANCORA_ANTERIOR")

            if delta > MAX_SALTO_ABAST:
                motivos.append("SALTO_ANTERIOR_MAIOR_1500")

            if pd.notna(km_cli):
                tolerancia = max(TOLERANCIA_ANCORA_ABS, km_cli * TOLERANCIA_ANCORA_PCT)
                if delta + tolerancia < km_cli:
                    motivos.append("ABAIXO_DISTANCIA_CLIENTE_ANTERIOR")
                else:
                    score += PESO_CLIENTE

    if idx_next is not None:
        od_next = grupo.loc[idx_next, "ODOMETRO_MATCH"]
        dt_next = grupo.loc[idx_next, "DATA_HORA_ABASTECIMENTO"]

        if pd.notna(od_next):
            delta = float(od_next) - od_cand
            km_cli = somar_distancia_producao(df_prod_placa, dt_abast, dt_next)

            if delta <= 1:
                motivos.append("NAO_EVOLUI_ATE_ANCORA_POSTERIOR")

            if delta > MAX_SALTO_ABAST:
                motivos.append("SALTO_POSTERIOR_MAIOR_1500")

            if pd.notna(km_cli):
                tolerancia = max(TOLERANCIA_ANCORA_ABS, km_cli * TOLERANCIA_ANCORA_PCT)
                if delta + tolerancia < km_cli:
                    motivos.append("ABAIXO_DISTANCIA_CLIENTE_POSTERIOR")
                else:
                    score += PESO_CLIENTE

    if idx_prev is not None and idx_next is not None:
        od_prev = grupo.loc[idx_prev, "ODOMETRO_MATCH"]
        od_next = grupo.loc[idx_next, "ODOMETRO_MATCH"]
        if pd.notna(od_prev) and pd.notna(od_next):
            if float(od_prev) < od_cand < float(od_next):
                score += PESO_CORREDOR
            else:
                motivos.append("FORA_CORREDOR_ENTRE_ANCORAS")

    ok = len(motivos) == 0
    return ok, score, "|".join(motivos) if motivos else None


def definir_ancoras(grupo, candidatos_local, df_prod_placa):
    grupo = grupo.sort_values("DATA_HORA_ABASTECIMENTO").copy().reset_index(drop=True)

    grupo["ODOMETRO_MATCH"] = np.nan
    grupo["MATCH_VALIDADO"] = False
    grupo["EH_ANCORA"] = False
    grupo["ORIGEM_ODOMETRO_MATCH"] = None
    grupo["CONFIANCA_ODOMETRO_MATCH"] = "SEM_ODOMETRO"
    grupo["MOTIVO_ODOMETRO_MATCH"] = None
    grupo["SCORE_MATCH"] = np.nan

    # Primeira passada: usa melhor candidato por horario/periodo
    for i in range(len(grupo)):
        cands = candidatos_local.get(i, pd.DataFrame())

        if cands is None or cands.empty:
            continue

        melhor = cands.sort_values(
            by=["_SCORE_INICIAL", "DIFERENCA_MINUTOS"],
            ascending=[False, True]
        ).iloc[0]

        grupo.loc[i, "ODOMETRO_MATCH"] = float(melhor["ODOMETRO"])
        grupo.loc[i, "MATCH_VALIDADO"] = True
        grupo.loc[i, "EH_ANCORA"] = True
        grupo.loc[i, "DATA_HORA_MAXTRACK_ENCONTRADA"] = melhor["DATA_HORA_REF"]
        grupo.loc[i, "TIPO_REF_MAXTRACK"] = melhor["TIPO_REF"]
        grupo.loc[i, "DIFERENCA_MINUTOS"] = float(melhor["DIFERENCA_MINUTOS"])
        grupo.loc[i, "CONFIANCA_TEMPORAL"] = classificar_confianca_temporal(float(melhor["DIFERENCA_MINUTOS"]))
        grupo.loc[i, "ORIGEM_ODOMETRO_MATCH"] = "MAXTRACK_DIRETO"
        grupo.loc[i, "CONFIANCA_ODOMETRO_MATCH"] = "ALTA"
        grupo.loc[i, "MOTIVO_ODOMETRO_MATCH"] = "ANCORA_MAXTRACK"
        grupo.loc[i, "SCORE_MATCH"] = float(melhor["_SCORE_INICIAL"])

    # Segunda passada: revalida com contexto para frente e para tras
    for _ in range(3):
        alterou = False

        for i in range(len(grupo)):
            cands = candidatos_local.get(i, pd.DataFrame())

            if cands is None or cands.empty:
                if grupo.loc[i, "EH_ANCORA"] == True:
                    grupo.loc[i, "ODOMETRO_MATCH"] = np.nan
                    grupo.loc[i, "MATCH_VALIDADO"] = False
                    grupo.loc[i, "EH_ANCORA"] = False
                    grupo.loc[i, "ORIGEM_ODOMETRO_MATCH"] = None
                    grupo.loc[i, "CONFIANCA_ODOMETRO_MATCH"] = "SEM_ODOMETRO"
                    grupo.loc[i, "MOTIVO_ODOMETRO_MATCH"] = "SEM_CANDIDATO_MAXTRACK"
                    alterou = True
                continue

            melhor_ok = None
            melhor_score = -999999
            melhor_motivo_rejeicao = None

            for _, cand in cands.iterrows():
                ok, score, motivo = avaliar_candidato_contexto(grupo, i, cand, df_prod_placa)
                if ok and score > melhor_score:
                    melhor_ok = cand
                    melhor_score = score
                    melhor_motivo_rejeicao = None
                elif melhor_ok is None:
                    melhor_motivo_rejeicao = motivo

            if melhor_ok is None:
                if grupo.loc[i, "EH_ANCORA"] == True:
                    grupo.loc[i, "ODOMETRO_MATCH"] = np.nan
                    grupo.loc[i, "MATCH_VALIDADO"] = False
                    grupo.loc[i, "EH_ANCORA"] = False
                    grupo.loc[i, "ORIGEM_ODOMETRO_MATCH"] = None
                    grupo.loc[i, "CONFIANCA_ODOMETRO_MATCH"] = "SEM_ODOMETRO"
                    grupo.loc[i, "MOTIVO_ODOMETRO_MATCH"] = melhor_motivo_rejeicao or "REJEITADO_CONTEXTO"
                    alterou = True
                continue

            novo_odom = float(melhor_ok["ODOMETRO"])
            od_atual = grupo.loc[i, "ODOMETRO_MATCH"]

            if pd.isna(od_atual) or abs(float(od_atual) - novo_odom) > 0:
                grupo.loc[i, "ODOMETRO_MATCH"] = novo_odom
                grupo.loc[i, "MATCH_VALIDADO"] = True
                grupo.loc[i, "EH_ANCORA"] = True
                grupo.loc[i, "DATA_HORA_MAXTRACK_ENCONTRADA"] = melhor_ok["DATA_HORA_REF"]
                grupo.loc[i, "TIPO_REF_MAXTRACK"] = melhor_ok["TIPO_REF"]
                grupo.loc[i, "DIFERENCA_MINUTOS"] = float(melhor_ok["DIFERENCA_MINUTOS"])
                grupo.loc[i, "CONFIANCA_TEMPORAL"] = classificar_confianca_temporal(float(melhor_ok["DIFERENCA_MINUTOS"]))
                grupo.loc[i, "ORIGEM_ODOMETRO_MATCH"] = "MAXTRACK_DIRETO"
                grupo.loc[i, "CONFIANCA_ODOMETRO_MATCH"] = "ALTA"
                grupo.loc[i, "MOTIVO_ODOMETRO_MATCH"] = "ANCORA_MAXTRACK_VALIDADA"
                grupo.loc[i, "SCORE_MATCH"] = melhor_score
                alterou = True

        if not alterou:
            break

    return grupo


# ============================================================
# PREENCHER ODOMETRO_MATCH
# ============================================================
def preencher_odometro_match(grupo, df_prod_placa):
    grupo = grupo.sort_values("DATA_HORA_ABASTECIMENTO").copy().reset_index(drop=True)

    idx_anchors = grupo.index[grupo["EH_ANCORA"] == True].tolist()

    if len(idx_anchors) == 0:
        grupo["ORIGEM_ODOMETRO_MATCH"] = grupo["ORIGEM_ODOMETRO_MATCH"].fillna("SEM_MAXTRACK")
        grupo["CONFIANCA_ODOMETRO_MATCH"] = grupo["CONFIANCA_ODOMETRO_MATCH"].fillna("MUITO_BAIXA")
        grupo["MOTIVO_ODOMETRO_MATCH"] = grupo["MOTIVO_ODOMETRO_MATCH"].fillna("SEM_ANCORA_MAXTRACK")
        return grupo

    # Antes da primeira ancora
    idx_first = idx_anchors[0]
    od_first = grupo.loc[idx_first, "ODOMETRO_MATCH"]
    dt_first = grupo.loc[idx_first, "DATA_HORA_ABASTECIMENTO"]

    for i in range(idx_first - 1, -1, -1):
        dt_i = grupo.loc[i, "DATA_HORA_ABASTECIMENTO"]
        dist = somar_distancia_producao(df_prod_placa, dt_i, dt_first)

        if pd.notna(dist):
            od_calc = float(od_first) - max(np.ceil(dist), MIN_INCREMENTO_ABSOLUTO)
            origem = "CLIENTE_ATE_PRIMEIRO_MAXTRACK"
            conf = "BAIXA"
            motivo = "PREENCHIDO_COM_CLIENTE_ANTES_PRIMEIRA_ANCORA"
        else:
            od_calc = float(od_first) - (MIN_INCREMENTO_SEM_CLIENTE * (idx_first - i))
            origem = "TENDENCIA_FALLBACK"
            conf = "MUITO_BAIXA"
            motivo = "PREENCHIDO_SEM_CLIENTE_ANTES_PRIMEIRA_ANCORA"

        grupo.loc[i, "ODOMETRO_MATCH"] = round(od_calc, 0)
        grupo.loc[i, "ORIGEM_ODOMETRO_MATCH"] = origem
        grupo.loc[i, "CONFIANCA_ODOMETRO_MATCH"] = conf
        grupo.loc[i, "MOTIVO_ODOMETRO_MATCH"] = motivo

    # Entre ancoras
    for k in range(len(idx_anchors) - 1):
        a = idx_anchors[k]
        b = idx_anchors[k + 1]

        od_a = grupo.loc[a, "ODOMETRO_MATCH"]
        od_b = grupo.loc[b, "ODOMETRO_MATCH"]
        dt_a = grupo.loc[a, "DATA_HORA_ABASTECIMENTO"]
        dt_b = grupo.loc[b, "DATA_HORA_ABASTECIMENTO"]

        if pd.isna(od_a) or pd.isna(od_b) or float(od_b) <= float(od_a):
            continue

        delta_total = float(od_b) - float(od_a)
        dist_total = somar_distancia_producao(df_prod_placa, dt_a, dt_b)
        usar_tempo = pd.isna(dist_total) or dist_total <= 0

        for j in range(a + 1, b):
            if grupo.loc[j, "EH_ANCORA"] == True:
                continue

            dt_j = grupo.loc[j, "DATA_HORA_ABASTECIMENTO"]

            if usar_tempo:
                dur_total = (dt_b - dt_a).total_seconds()
                dur_parcial = (dt_j - dt_a).total_seconds()

                if dur_total <= 0:
                    continue

                proporcao = max(0, min(1, dur_parcial / dur_total))
                od_calc = float(od_a) + delta_total * proporcao
                origem = "TEMPO_ENTRE_MAXTRACK"
                conf = "BAIXA"
                motivo = "PREENCHIDO_POR_TEMPO_ENTRE_ANCORAS"

            else:
                dist_parcial = somar_distancia_producao(df_prod_placa, dt_a, dt_j)
                dist_restante = somar_distancia_producao(df_prod_placa, dt_j, dt_b)

                if pd.isna(dist_parcial):
                    continue

                min_possivel = float(od_a) + max(np.ceil(dist_parcial), 0)

                if pd.notna(dist_restante):
                    max_possivel = float(od_b) - max(np.ceil(dist_restante), 0)
                else:
                    max_possivel = float(od_b)

                proporcao = max(0, min(1, dist_parcial / dist_total))
                od_estimado = float(od_a) + delta_total * proporcao
                od_calc = min(max(od_estimado, min_possivel), max_possivel)

                origem = "CLIENTE_ENTRE_MAXTRACK"
                conf = "MEDIA"
                motivo = "PREENCHIDO_COM_CLIENTE_ENTRE_ANCORAS"

            grupo.loc[j, "ODOMETRO_MATCH"] = round(od_calc, 0)
            grupo.loc[j, "ORIGEM_ODOMETRO_MATCH"] = origem
            grupo.loc[j, "CONFIANCA_ODOMETRO_MATCH"] = conf
            grupo.loc[j, "MOTIVO_ODOMETRO_MATCH"] = motivo

    # Depois da ultima ancora
    idx_last = idx_anchors[-1]
    od_last = grupo.loc[idx_last, "ODOMETRO_MATCH"]
    dt_last = grupo.loc[idx_last, "DATA_HORA_ABASTECIMENTO"]

    for i in range(idx_last + 1, len(grupo)):
        if grupo.loc[i, "EH_ANCORA"] == True:
            continue

        dt_i = grupo.loc[i, "DATA_HORA_ABASTECIMENTO"]
        dist = somar_distancia_producao(df_prod_placa, dt_last, dt_i)

        if pd.notna(dist):
            od_calc = float(od_last) + max(np.ceil(dist), MIN_INCREMENTO_ABSOLUTO)
            origem = "CLIENTE_APOS_ULTIMO_MAXTRACK"
            conf = "BAIXA"
            motivo = "PREENCHIDO_COM_CLIENTE_APOS_ULTIMA_ANCORA"
        else:
            od_calc = float(od_last) + (MIN_INCREMENTO_SEM_CLIENTE * (i - idx_last))
            origem = "TENDENCIA_FALLBACK"
            conf = "MUITO_BAIXA"
            motivo = "PREENCHIDO_SEM_CLIENTE_APOS_ULTIMA_ANCORA"

        grupo.loc[i, "ODOMETRO_MATCH"] = round(od_calc, 0)
        grupo.loc[i, "ORIGEM_ODOMETRO_MATCH"] = origem
        grupo.loc[i, "CONFIANCA_ODOMETRO_MATCH"] = conf
        grupo.loc[i, "MOTIVO_ODOMETRO_MATCH"] = motivo

    return grupo


# ============================================================
# AJUSTES FINAIS NO PROPRIO ODOMETRO_MATCH
# ============================================================
def ajustar_odometro_match_final(grupo, df_prod_placa):
    grupo = grupo.sort_values("DATA_HORA_ABASTECIMENTO").copy().reset_index(drop=True)

    grupo["KM_CLIENTE_INTERVALO"] = np.nan
    grupo["KM_ENTRE_ABAST"] = np.nan
    grupo["DIFERENCA_ODOM_X_CLIENTE"] = np.nan
    grupo["FLAG_INTERVALO_FORA_FAIXA"] = "NAO"
    grupo["AJUSTE_ODOMETRO_MATCH"] = "NAO"

    for i in range(1, len(grupo)):
        od_prev = grupo.loc[i - 1, "ODOMETRO_MATCH"]
        od_cur = grupo.loc[i, "ODOMETRO_MATCH"]
        dt_prev = grupo.loc[i - 1, "DATA_HORA_ABASTECIMENTO"]
        dt_cur = grupo.loc[i, "DATA_HORA_ABASTECIMENTO"]

        if pd.isna(od_prev) or pd.isna(od_cur):
            continue

        km_cliente = somar_distancia_producao(df_prod_placa, dt_prev, dt_cur)
        grupo.loc[i, "KM_CLIENTE_INTERVALO"] = km_cliente

        km_odom = float(od_cur) - float(od_prev)

        if pd.notna(km_cliente):
            min_ok = np.ceil(km_cliente + MARGEM_TECNICA_CLIENTE)
            max_ok = MAX_SALTO_ABAST
        else:
            tendencia = mediana_km_validos(grupo, i, JANELA_TENDENCIA)
            if pd.isna(tendencia):
                min_ok = MIN_INCREMENTO_SEM_CLIENTE
                max_ok = min(MAX_VARIACAO_LOCAL_SEM_CLIENTE, MAX_SALTO_ABAST)
            else:
                min_ok = max(MIN_INCREMENTO_SEM_CLIENTE, np.floor(tendencia * 0.8))
                max_ok = min(MAX_VARIACAO_LOCAL_SEM_CLIENTE, np.ceil(tendencia * 1.4 + 40), MAX_SALTO_ABAST)

        ajuste = "NAO"

        if km_odom <= 1:
            novo = float(od_prev) + max(min_ok, MIN_INCREMENTO_ABSOLUTO)
            grupo.loc[i, "ODOMETRO_MATCH"] = round(novo, 0)
            ajuste = "AJUSTADO_KM_MENOR_OU_IGUAL_1"

        elif km_odom < min_ok:
            novo = float(od_prev) + min_ok
            grupo.loc[i, "ODOMETRO_MATCH"] = round(novo, 0)
            ajuste = "AJUSTADO_PARA_DISTANCIA_CLIENTE"

        elif km_odom > max_ok:
            novo = float(od_prev) + max_ok
            grupo.loc[i, "ODOMETRO_MATCH"] = round(novo, 0)
            ajuste = "AJUSTADO_SALTO_MAIOR_1500"

        if ajuste != "NAO":
            grupo.loc[i, "AJUSTE_ODOMETRO_MATCH"] = ajuste
            grupo.loc[i, "FLAG_INTERVALO_FORA_FAIXA"] = "SIM"

            motivo_atual = grupo.loc[i, "MOTIVO_ODOMETRO_MATCH"]
            if pd.isna(motivo_atual) or motivo_atual is None:
                motivo_atual = ""
            grupo.loc[i, "MOTIVO_ODOMETRO_MATCH"] = (str(motivo_atual) + "|" + ajuste).strip("|")

            if grupo.loc[i, "EH_ANCORA"] == True:
                grupo.loc[i, "CONFIANCA_ODOMETRO_MATCH"] = "MEDIA"
                grupo.loc[i, "ORIGEM_ODOMETRO_MATCH"] = "MAXTRACK_AJUSTADO_REGRA"
            else:
                grupo.loc[i, "CONFIANCA_ODOMETRO_MATCH"] = "BAIXA"

        # recalcula depois do ajuste
        od_cur2 = grupo.loc[i, "ODOMETRO_MATCH"]
        if pd.notna(od_cur2):
            km_final = float(od_cur2) - float(od_prev)
            grupo.loc[i, "KM_ENTRE_ABAST"] = km_final
            if pd.notna(km_cliente):
                grupo.loc[i, "DIFERENCA_ODOM_X_CLIENTE"] = km_final - km_cliente
                if km_final < km_cliente or km_final > MAX_SALTO_ABAST:
                    grupo.loc[i, "FLAG_INTERVALO_FORA_FAIXA"] = "SIM"
            else:
                if km_final <= 1 or km_final > MAX_SALTO_ABAST:
                    grupo.loc[i, "FLAG_INTERVALO_FORA_FAIXA"] = "SIM"

    return grupo


# ============================================================
# RESUMOS
# ============================================================
def gerar_resumo_mensal(resultado_final, prod):
    base = resultado_final.copy()
    base = base[base["ODOMETRO_MATCH"].notna()].copy()
    base["ANO_MES"] = base["DATA_HORA_ABASTECIMENTO"].dt.to_period("M").astype(str)

    resumo_odom = (
        base.sort_values(["PLACA", "DATA_HORA_ABASTECIMENTO"])
        .groupby(["PLACA", "ANO_MES"], as_index=False)
        .agg(
            PRIMEIRA_DATA_MES=("DATA_HORA_ABASTECIMENTO", "min"),
            ULTIMA_DATA_MES=("DATA_HORA_ABASTECIMENTO", "max"),
            PRIMEIRO_ODOMETRO_MES=("ODOMETRO_MATCH", "first"),
            ULTIMO_ODOMETRO_MES=("ODOMETRO_MATCH", "last"),
            QTD_ABAST_MES=("ODOMETRO_MATCH", "size")
        )
    )

    resumo_odom["KM_RODADO_ODOMETRO_MES"] = (
        resumo_odom["ULTIMO_ODOMETRO_MES"] - resumo_odom["PRIMEIRO_ODOMETRO_MES"]
    )

    prod_mes = prod.copy()
    prod_mes["ANO_MES"] = prod_mes["DATA_HORA_PROD"].dt.to_period("M").astype(str)

    resumo_prod = (
        prod_mes.groupby(["PLACA", "ANO_MES"], as_index=False)
        .agg(
            DISTANCIA_CLIENTE_MES=("DISTANCIA_PROD", "sum"),
            PRIMEIRA_DATA_CLIENTE_MES=("DATA_HORA_PROD", "min"),
            ULTIMA_DATA_CLIENTE_MES=("DATA_HORA_PROD", "max"),
            QTD_EVENTOS_CLIENTE_MES=("DISTANCIA_PROD", "size")
        )
    )

    resumo = resumo_odom.merge(resumo_prod, on=["PLACA", "ANO_MES"], how="outer")
    resumo["DIFERENCA_ODOM_X_CLIENTE_MES"] = resumo["KM_RODADO_ODOMETRO_MES"] - resumo["DISTANCIA_CLIENTE_MES"]
    resumo["DIFERENCA_ABS_MES"] = resumo["DIFERENCA_ODOM_X_CLIENTE_MES"].abs()
    resumo["CLASSIFICACAO_DIFERENCA_MES"] = resumo["DIFERENCA_ODOM_X_CLIENTE_MES"].apply(classificar_diff_mes)

    return resumo.sort_values(["PLACA", "ANO_MES"]).reset_index(drop=True)


# ============================================================
# FORMATACAO EXCEL
# ============================================================
def formatar_excel(caminho):
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = load_workbook(caminho)

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            max_col = ws.max_column
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 10
                for cell in ws[col_letter][:200]:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

        wb.save(caminho)
    except Exception as e:
        print(f"Arquivo gerado, mas nao foi possivel aplicar formatacao Excel: {e}")


# ============================================================
# MAIN
# ============================================================
def main():
    arq_combustivel, arq_maxtrack, arq_ativo, arq_producao, arq_saida = selecionar_bases()

    print("\nLendo planilhas...")

    df_comb = ler_excel_com_aba(arq_combustivel, ABA_COMBUSTIVEL)
    df_max = ler_excel_com_aba(arq_maxtrack, ABA_MAXTRACK)
    df_ativo = ler_excel_com_aba(arq_ativo, ABA_ATIVO)
    df_prod = ler_excel_com_aba(arq_producao, ABA_PRODUCAO)

    print("Leitura concluida.")

    # ========================================================
    # IDENTIFICACAO DE COLUNAS
    # ========================================================
    print("Identificando colunas...")

    col_comb_veiculo = escolher_coluna(df_comb, ["Veiculo", "Veiculo ", "Veículo"])
    col_comb_data = escolher_coluna(df_comb, ["Data", "Data Ajustada", "Data Abastecimento"])
    col_comb_litros = escolher_coluna(df_comb, ["Volume (L)", "Litros", "Volume"])

    col_ativo_go = escolher_coluna(df_ativo, ["GO", "G.O", "CODIGO GO", "CODIGO"])
    try:
        col_ativo_placa = escolher_coluna(df_ativo, ["PLACA", "Placa", "Placa do Veiculo", "Placa do Veículo"])
    except KeyError:
        col_ativo_placa = escolher_melhor_coluna_por_palavras(df_ativo, ["PLACA"], obrigatoria=True)

    try:
        col_max_placa = escolher_coluna(df_max, ["Identificador/Placa", "Identificador / Placa", "Placa", "Identificador"])
    except KeyError:
        col_max_placa = escolher_melhor_coluna_por_palavras(df_max, ["PLACA"], obrigatoria=True)

    col_max_inicio = escolher_coluna(df_max, ["Início", "Inicio"], obrigatoria=False)
    col_max_fim = escolher_coluna(df_max, ["Fim"], obrigatoria=False)
    col_max_odom_ini = escolher_coluna(df_max, ["Odômetro Inicial", "Odometro Inicial"], obrigatoria=False)
    col_max_odom_fim = escolher_coluna(df_max, ["Odômetro Final", "Odometro Final"], obrigatoria=False)

    try:
        col_prod_placa = escolher_coluna(df_prod, ["Placa"])
    except KeyError:
        col_prod_placa = escolher_melhor_coluna_por_palavras(df_prod, ["PLACA"], obrigatoria=True)

    col_prod_saida = escolher_coluna(df_prod, ["Saída", "Saida", "Hora Saída", "Hora Saida", "Data Saida"])
    col_prod_dist = escolher_coluna(df_prod, ["Distância", "Distancia", "Distancia KM", "KM"])

    print("Colunas identificadas com sucesso.")

    # ========================================================
    # TRATAMENTO ATIVO
    # ========================================================
    print("Tratando base Ativo...")

    ativo = df_ativo[[col_ativo_go, col_ativo_placa]].copy()
    ativo.columns = ["GO", "PLACA"]
    ativo["GO"] = ativo["GO"].apply(lambda x: "".join(ch for ch in str(x) if ch.isdigit()) if pd.notna(x) else None)
    ativo["PLACA"] = ativo["PLACA"].apply(normalizar_placa)
    ativo = ativo.dropna(subset=["GO", "PLACA"]).drop_duplicates()

    print(f"Ativo tratado: {len(ativo)} registros.")

    # ========================================================
    # TRATAMENTO COMBUSTIVEL
    # ========================================================
    print("Tratando base Combustivel...")

    comb = df_comb[[col_comb_veiculo, col_comb_data, col_comb_litros]].copy()
    comb.columns = ["VEICULO", "DATA_HORA_ABASTECIMENTO", "LITROS"]
    comb["GO"] = comb["VEICULO"].apply(extrair_go_do_veiculo)
    comb["DATA_HORA_ABASTECIMENTO"] = pd.to_datetime(comb["DATA_HORA_ABASTECIMENTO"], errors="coerce")
    comb["LITROS"] = comb["LITROS"].apply(para_numero)

    comb = comb.merge(ativo, on="GO", how="left")
    comb["PLACA"] = comb["PLACA"].apply(normalizar_placa)
    comb = comb.dropna(subset=["DATA_HORA_ABASTECIMENTO"]).copy()
    comb = comb.sort_values(["PLACA", "DATA_HORA_ABASTECIMENTO"], na_position="last").reset_index(drop=True)
    comb["ID_LINHA_ORIGINAL"] = comb.index.astype(str)

    print(f"Combustivel tratado: {len(comb)} registros.")

    # ========================================================
    # TRATAMENTO MAXTRACK
    # ========================================================
    print("Tratando base Maxtrack...")

    eventos = []

    if col_max_inicio and col_max_odom_ini:
        tmp_ini = df_max[[col_max_placa, col_max_inicio, col_max_odom_ini]].copy()
        tmp_ini.columns = ["PLACA", "DATA_HORA_REF", "ODOMETRO"]
        tmp_ini["TIPO_REF"] = "INICIO"
        eventos.append(tmp_ini)

    if col_max_fim and col_max_odom_fim:
        tmp_fim = df_max[[col_max_placa, col_max_fim, col_max_odom_fim]].copy()
        tmp_fim.columns = ["PLACA", "DATA_HORA_REF", "ODOMETRO"]
        tmp_fim["TIPO_REF"] = "FIM"
        eventos.append(tmp_fim)

    if not eventos:
        raise ValueError("Nao foi possivel montar a base Maxtrack. Verifique colunas de inicio/fim e odometro.")

    odom = pd.concat(eventos, ignore_index=True)
    odom["PLACA"] = odom["PLACA"].apply(normalizar_placa)
    odom["DATA_HORA_REF"] = pd.to_datetime(odom["DATA_HORA_REF"], errors="coerce")
    odom["ODOMETRO"] = odom["ODOMETRO"].apply(para_numero)
    odom = odom.dropna(subset=["PLACA", "DATA_HORA_REF", "ODOMETRO"]).copy()
    odom = odom[odom["ODOMETRO"] > 0].copy()
    odom = odom.drop_duplicates(subset=["PLACA", "DATA_HORA_REF", "ODOMETRO", "TIPO_REF"])
    odom = odom.sort_values(["PLACA", "DATA_HORA_REF"]).reset_index(drop=True)

    placa_para_eventos = {
        placa: grupo.sort_values("DATA_HORA_REF").reset_index(drop=True)
        for placa, grupo in odom.groupby("PLACA")
    }

    print(f"Maxtrack tratado: {len(odom)} eventos.")

    # ========================================================
    # TRATAMENTO PRODUCAO
    # ========================================================
    print("Tratando base Cliente / Producao...")

    prod = df_prod[[col_prod_placa, col_prod_saida, col_prod_dist]].copy()
    prod.columns = ["PLACA", "DATA_HORA_PROD", "DISTANCIA_PROD"]
    prod["PLACA"] = prod["PLACA"].apply(normalizar_placa)
    prod["DATA_HORA_PROD"] = pd.to_datetime(prod["DATA_HORA_PROD"], errors="coerce")
    prod["DISTANCIA_PROD"] = prod["DISTANCIA_PROD"].apply(para_numero)
    prod = prod.dropna(subset=["PLACA", "DATA_HORA_PROD", "DISTANCIA_PROD"]).copy()
    prod = prod[prod["DISTANCIA_PROD"] >= 0].copy()
    prod = prod.sort_values(["PLACA", "DATA_HORA_PROD"]).reset_index(drop=True)

    placa_para_prod = {
        placa: grupo.sort_values("DATA_HORA_PROD").reset_index(drop=True)
        for placa, grupo in prod.groupby("PLACA")
    }

    print(f"Cliente / Producao tratada: {len(prod)} registros.")

    # ========================================================
    # MATCH INICIAL
    # ========================================================
    imprimir_etapa("Buscando candidatos Maxtrack por abastecimento")

    resultados = []
    total_comb = len(comb)

    for i, row in comb.iterrows():
        placa = row["PLACA"]
        dt_abast = row["DATA_HORA_ABASTECIMENTO"]

        base_linha = {
            "PLACA": placa,
            "GO": row["GO"],
            "VEICULO": row["VEICULO"],
            "ID_LINHA_ORIGINAL": row["ID_LINHA_ORIGINAL"],
            "DATA_HORA_ABASTECIMENTO": dt_abast,
            "LITROS": row["LITROS"],
            "ODOMETRO_CANDIDATO": np.nan,
            "DATA_HORA_MAXTRACK_ENCONTRADA": pd.NaT,
            "TIPO_REF_MAXTRACK": None,
            "DIFERENCA_MINUTOS": np.nan,
            "STATUS_MATCH": "NAO_ENCONTRADO",
            "CONFIANCA_TEMPORAL": "NENHUM",
        }

        if not placa:
            base_linha["STATUS_MATCH"] = "SEM_PLACA"
            resultados.append(base_linha)
            continue

        df_placa = placa_para_eventos.get(placa, pd.DataFrame())
        cands = listar_candidatos_match(df_placa, dt_abast)

        if not cands.empty:
            melhor = cands.iloc[0]
            base_linha["ODOMETRO_CANDIDATO"] = float(melhor["ODOMETRO"])
            base_linha["DATA_HORA_MAXTRACK_ENCONTRADA"] = melhor["DATA_HORA_REF"]
            base_linha["TIPO_REF_MAXTRACK"] = melhor["TIPO_REF"]
            base_linha["DIFERENCA_MINUTOS"] = round(float(melhor["DIFERENCA_MINUTOS"]), 2)
            base_linha["STATUS_MATCH"] = "ENCONTRADO"
            base_linha["CONFIANCA_TEMPORAL"] = classificar_confianca_temporal(float(melhor["DIFERENCA_MINUTOS"]))

        resultados.append(base_linha)

        if (i + 1) % 50 == 0 or (i + 1) == total_comb:
            mostrar_progresso("Candidatos iniciais", i + 1, total_comb)

    resultado_match = pd.DataFrame(resultados)
    resultado_match = resultado_match.sort_values(["PLACA", "DATA_HORA_ABASTECIMENTO"], na_position="last").reset_index(drop=True)

    print("Candidatos iniciais concluidos.")

    # ========================================================
    # CONSOLIDACAO
    # ========================================================
    imprimir_etapa("Consolidando abastecimentos proximos")

    agrupados = []
    grupos_match = list(resultado_match.groupby("PLACA", dropna=False))
    total_grupos_match = len(grupos_match)

    for pos, (placa, grupo) in enumerate(grupos_match, start=1):
        agrupados.append(agrupar_abastecimentos_proximos(grupo))
        if pos % 10 == 0 or pos == total_grupos_match:
            mostrar_progresso("Consolidacao por placa", pos, total_grupos_match)

    resultado_consolidado = pd.concat(agrupados, ignore_index=True)
    resultado_consolidado = resultado_consolidado.sort_values(["PLACA", "DATA_HORA_ABASTECIMENTO"], na_position="last").reset_index(drop=True)

    print(f"Abastecimentos apos consolidacao: {len(resultado_consolidado)}")

    # ========================================================
    # DEFINIR ANCORAS POR PLACA
    # ========================================================
    imprimir_etapa("Definindo ancoras Maxtrack validas")

    validado = []
    grupos_consolidado = list(resultado_consolidado.groupby("PLACA", dropna=False))
    total_grupos_consolidado = len(grupos_consolidado)

    for pos, (placa, grupo) in enumerate(grupos_consolidado, start=1):
        grupo = grupo.copy().reset_index(drop=True)

        candidatos_local = {}
        for idx in grupo.index:
            placa_local = grupo.loc[idx, "PLACA"]
            dt_local = grupo.loc[idx, "DATA_HORA_ABASTECIMENTO"]

            if not placa_local:
                candidatos_local[idx] = pd.DataFrame()
            else:
                candidatos_local[idx] = listar_candidatos_match(
                    placa_para_eventos.get(placa_local, pd.DataFrame()),
                    dt_local
                )

        df_prod_placa = placa_para_prod.get(placa)
        validado.append(definir_ancoras(grupo, candidatos_local, df_prod_placa))

        if pos % 5 == 0 or pos == total_grupos_consolidado:
            mostrar_progresso("Ancoras por placa", pos, total_grupos_consolidado)

    resultado_validado = pd.concat(validado, ignore_index=True)
    resultado_validado = resultado_validado.sort_values(["PLACA", "DATA_HORA_ABASTECIMENTO"], na_position="last").reset_index(drop=True)

    print("Ancoras definidas.")

    # ========================================================
    # PREENCHER ODOMETRO_MATCH
    # ========================================================
    imprimir_etapa("Preenchendo ODOMETRO_MATCH unico")

    preenchidos = []
    grupos_validado = list(resultado_validado.groupby("PLACA", dropna=False))
    total_grupos_validado = len(grupos_validado)

    for pos, (placa, grupo) in enumerate(grupos_validado, start=1):
        df_prod_placa = placa_para_prod.get(placa)
        preenchidos.append(preencher_odometro_match(grupo, df_prod_placa))

        if pos % 10 == 0 or pos == total_grupos_validado:
            mostrar_progresso("Preenchimento por placa", pos, total_grupos_validado)

    resultado_preenchido = pd.concat(preenchidos, ignore_index=True)
    resultado_preenchido = resultado_preenchido.sort_values(["PLACA", "DATA_HORA_ABASTECIMENTO"], na_position="last").reset_index(drop=True)

    print("ODOMETRO_MATCH preenchido.")

    # ========================================================
    # AJUSTE FINAL
    # ========================================================
    imprimir_etapa("Aplicando regras finais no ODOMETRO_MATCH")

    finais = []
    grupos_preenchido = list(resultado_preenchido.groupby("PLACA", dropna=False))
    total_grupos_preenchido = len(grupos_preenchido)

    for pos, (placa, grupo) in enumerate(grupos_preenchido, start=1):
        df_prod_placa = placa_para_prod.get(placa)
        finais.append(ajustar_odometro_match_final(grupo, df_prod_placa))

        if pos % 10 == 0 or pos == total_grupos_preenchido:
            mostrar_progresso("Ajuste final por placa", pos, total_grupos_preenchido)

    resultado_final = pd.concat(finais, ignore_index=True)
    resultado_final = resultado_final.sort_values(["PLACA", "DATA_HORA_ABASTECIMENTO"], na_position="last").reset_index(drop=True)

    print("Regras finais aplicadas.")

    # ========================================================
    # RESUMOS
    # ========================================================
    imprimir_etapa("Gerando resumo mensal e indicadores")

    resumo_mensal = gerar_resumo_mensal(resultado_final, prod)

    indicadores = pd.DataFrame({
        "Indicador": [
            "Total abastecimentos originais",
            "Total apos consolidacao",
            "Eventos agrupados",
            "ODOMETRO_MATCH preenchido",
            "ODOMETRO_MATCH vazio",
            "Matches Maxtrack validados",
            "Anchors Maxtrack",
            "Intervalos fora da faixa",
            "Saltos > 1500",
            "Intervalos abaixo da distancia cliente",
            "Ajustes no ODOMETRO_MATCH"
        ],
        "Valor": [
            len(comb),
            len(resultado_final),
            int((resultado_final["AGRUPADO_DUPLICADO"] == "SIM").sum()),
            int(resultado_final["ODOMETRO_MATCH"].notna().sum()),
            int(resultado_final["ODOMETRO_MATCH"].isna().sum()),
            int((resultado_final["MATCH_VALIDADO"] == True).sum()),
            int((resultado_final["EH_ANCORA"] == True).sum()),
            int((resultado_final["FLAG_INTERVALO_FORA_FAIXA"] == "SIM").sum()),
            int((resultado_final["KM_ENTRE_ABAST"] > MAX_SALTO_ABAST).sum()),
            int(((resultado_final["DIFERENCA_ODOM_X_CLIENTE"] < 0) & resultado_final["DIFERENCA_ODOM_X_CLIENTE"].notna()).sum()),
            int((resultado_final["AJUSTE_ODOMETRO_MATCH"] != "NAO").sum())
        ]
    })

    # ========================================================
    # SAIDAS
    # ========================================================
    resumo = resultado_final[
        [
            "PLACA",
            "DATA_HORA_ABASTECIMENTO",
            "LITROS",
            "ODOMETRO_MATCH",
            "KM_ENTRE_ABAST",
            "KM_CLIENTE_INTERVALO",
            "DIFERENCA_ODOM_X_CLIENTE",
            "ORIGEM_ODOMETRO_MATCH",
            "CONFIANCA_ODOMETRO_MATCH",
            "MOTIVO_ODOMETRO_MATCH",
            "AJUSTE_ODOMETRO_MATCH",
            "FLAG_INTERVALO_FORA_FAIXA",
            "MATCH_VALIDADO",
            "EH_ANCORA",
            "STATUS_MATCH",
            "CONFIANCA_TEMPORAL",
            "AGRUPADO_DUPLICADO",
            "QTD_EVENTOS_GRUPO",
            "DATA_HORA_MAXTRACK_ENCONTRADA",
            "TIPO_REF_MAXTRACK",
            "DIFERENCA_MINUTOS",
            "SCORE_MATCH",
            "GO",
            "VEICULO",
            "GRUPO_ABASTECIMENTO",
            "IDS_LINHAS_ORIGEM"
        ]
    ].copy()

    imprimir_etapa("Exportando arquivo final")

    with pd.ExcelWriter(arq_saida, engine="openpyxl") as writer:
        resumo.to_excel(writer, sheet_name="Resumo", index=False)
        resultado_final.to_excel(writer, sheet_name="Base_Completa", index=False)
        resumo_mensal.to_excel(writer, sheet_name="Resumo_Mensal", index=False)
        indicadores.to_excel(writer, sheet_name="Indicadores", index=False)
        comb.to_excel(writer, sheet_name="Comb_Tratado", index=False)
        odom.to_excel(writer, sheet_name="Eventos_Maxtrack", index=False)
        prod.to_excel(writer, sheet_name="Producao_Tratada", index=False)

    formatar_excel(arq_saida)

    print("\n" + "=" * 70)
    print("Arquivo gerado com sucesso!")
    print(arq_saida)
    print("=" * 70)


if __name__ == "__main__":
    try:
        if len(sys.argv) >= 6:
            arq_combustivel = sys.argv[1]
            arq_maxtrack = sys.argv[2]
            arq_ativo = sys.argv[3]
            arq_producao = sys.argv[4]
            arq_saida = sys.argv[5]

            def selecionar_bases():
                return arq_combustivel, arq_maxtrack, arq_ativo, arq_producao, arq_saida

            main()
        else:
            main()

    except Exception as e:
        print("\nERRO NA EXECUCAO:")
        print(str(e))
        print("\nVerifique se os arquivos selecionados estao corretos e se as abas/colunas existem.")
        raise
