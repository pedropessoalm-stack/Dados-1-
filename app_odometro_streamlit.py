import os
import sys
import re
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta

import numpy as np
import pandas as pd

import streamlit as st
import tempfile
import traceback
from pathlib import Path

# ============================================================
# V15 - ODOMETRO_MATCH UNICO + MAXTRACK PRIORITARIO + CLIENTE CHAVEADO
# ============================================================
# Regras principais:
# - ODOMETRO_MATCH e a unica coluna oficial de odometro.
# - Maxtrack e a fonte principal quando houver match temporal confiavel.
# - Cliente/Producao complementa lacunas ate voltar ao Maxtrack.
# - Se Cliente no dia > 700 km, trata como possivel mais de uma viagem e chaveia
#   a contribuicao para evitar salto irreal de odometro.
# - Nenhum intervalo final deve ficar com KM_ENTRE_ABAST > 1500.
# - KM_ENTRE_ABAST deve ser >= distancia Cliente ajustada/chaveada no intervalo.
# - AM/PM e usado somente internamente no score, nao e exportado como coluna.
# ============================================================

ABA_COMBUSTIVEL = "Abastecimentos"
ABA_MAXTRACK = "RL - Viagens"
ABA_ATIVO = "Ativo atualizado"
ABA_PRODUCAO = "BD Transporte"

MINUTOS_AGRUPAMENTO = 30
MAX_DIFF_ODOM_AGRUPAMENTO = 160
JANELA_MATCH_MINUTOS = 360
TOP_CANDIDATOS_POR_ABAST = 10

MAX_SALTO_ABAST = 1500
LIMITE_CLIENTE_DIA_ALTO = 700
MIN_INCREMENTO_ABSOLUTO = 2
MIN_INCREMENTO_SEM_CLIENTE = 20
MAX_VARIACAO_LOCAL_SEM_CLIENTE = 800
JANELA_TENDENCIA = 3
MARGEM_TECNICA_CLIENTE = 0

PESO_MESMO_DIA = 3.0
PESO_MESMO_PERIODO = 3.0
PESO_TEMPO = 2.0
PESO_TIPO_REF = 0.8
PESO_CLIENTE = 3.0


# ============================================================
# PROGRESSO TERMINAL + STREAMLIT
# ============================================================
ST_PROGRESS_BAR = None
ST_STATUS_BOX = None
ST_LOG_BOX = None
ST_PROGRESS_VALUE = 0
ST_LOGS = []

ST_STAGE_RANGES = {
    "Leitura rapida Maxtrack": (12, 28),
    "Match Maxtrack por placa": (38, 58),
    "Consolidacao por placa": (58, 68),
    "Processamento por placa": (68, 88),
}

def configurar_progresso_streamlit(progress_bar=None, status_box=None, log_box=None):
    global ST_PROGRESS_BAR, ST_STATUS_BOX, ST_LOG_BOX, ST_PROGRESS_VALUE, ST_LOGS
    ST_PROGRESS_BAR = progress_bar
    ST_STATUS_BOX = status_box
    ST_LOG_BOX = log_box
    ST_PROGRESS_VALUE = 0
    ST_LOGS = []

def atualizar_progresso_streamlit(valor, texto):
    global ST_PROGRESS_VALUE, ST_LOGS
    valor = int(max(0, min(100, valor)))
    ST_PROGRESS_VALUE = max(ST_PROGRESS_VALUE, valor)
    if ST_PROGRESS_BAR is not None:
        ST_PROGRESS_BAR.progress(ST_PROGRESS_VALUE / 100.0, text=f"{ST_PROGRESS_VALUE}% - {texto}")
    if ST_STATUS_BOX is not None:
        ST_STATUS_BOX.info(f"{ST_PROGRESS_VALUE}% - {texto}")

def registrar_log_streamlit(texto):
    ST_LOGS.append(texto)
    if len(ST_LOGS) > 18:
        del ST_LOGS[:-18]
    if ST_LOG_BOX is not None:
        ST_LOG_BOX.code("\n".join(ST_LOGS), language="text")

def mostrar_progresso(etapa, atual, total):
    if total <= 0:
        return
    pct = (atual / total) * 100
    barra_total = 30
    preenchido = int(barra_total * atual / total)
    barra = "#" * preenchido + "-" * (barra_total - preenchido)
    print(f"\r{etapa}: [{barra}] {pct:6.2f}% ({atual}/{total})", end="", flush=True)
    if etapa in ST_STAGE_RANGES:
        ini, fim = ST_STAGE_RANGES[etapa]
        progresso = ini + ((fim - ini) * atual / total)
        atualizar_progresso_streamlit(progresso, f"{etapa}: {pct:0.2f}% ({atual}/{total})")
    if atual >= total:
        print()
        registrar_log_streamlit(f"OK - {etapa} concluido ({total}/{total})")


def imprimir_etapa(texto):
    print("\n" + "=" * 70)
    print(texto)
    print("=" * 70)
    registrar_log_streamlit(texto)
    mapa = {
        "Lendo e tratando bases": 5,
        "Preenchendo ODOMETRO_MATCH e aplicando regras finais": 68,
        "Gerando resumos e exportando": 88,
    }
    if texto in mapa:
        atualizar_progresso_streamlit(mapa[texto], texto)


def selecionar_arquivo_gui(titulo, descricao):
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        messagebox.showinfo("Selecionar arquivo", descricao)
        caminho = filedialog.askopenfilename(
            title=titulo,
            filetypes=[("Arquivos Excel", "*.xlsx *.xls"), ("Todos os arquivos", "*.*")],
        )
        root.destroy()
        if not caminho:
            print("Nenhum arquivo selecionado. Execucao cancelada.")
            sys.exit(0)
        return caminho
    except Exception:
        print("\n" + "=" * 70)
        print(descricao)
        print("=" * 70)
        caminho = input("Cole o caminho completo do arquivo Excel: ").strip().strip('"')
        if not caminho:
            print("Nenhum arquivo informado. Execucao cancelada.")
            sys.exit(0)
        return caminho


def selecionar_saida_gui(sugestao):
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        messagebox.showinfo(
            "Salvar resultado",
            "Agora escolha onde salvar o arquivo final gerado.\n\n"
            "Sugestao: Resumo_Abastecimento_Odometro_V15.xlsx",
        )
        caminho = filedialog.asksaveasfilename(
            title="Salvar arquivo final",
            initialfile=os.path.basename(sugestao),
            initialdir=os.path.dirname(sugestao) or os.getcwd(),
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
        )
        root.destroy()
        if not caminho:
            print("Nenhum local de saida selecionado. Execucao cancelada.")
            sys.exit(0)
        return caminho
    except Exception:
        print("\nEscolha onde salvar o arquivo final.")
        caminho = input(f"Caminho de saida sugerido [{sugestao}]: ").strip().strip('"')
        return caminho if caminho else os.path.abspath(sugestao)


def selecionar_bases():
    print("\n" + "=" * 70)
    print("SELECAO DAS BASES")
    print("=" * 70)
    arq_combustivel = selecionar_arquivo_gui(
        "1 de 4 - Selecione a BASE DE COMBUSTIVEL",
        f"1 de 4\n\nSelecione a BASE DE COMBUSTIVEL.\nAba esperada: {ABA_COMBUSTIVEL}",
    )
    arq_maxtrack = selecionar_arquivo_gui(
        "2 de 4 - Selecione a BASE MAXTRACK / KM RODADO",
        f"2 de 4\n\nSelecione a BASE MAXTRACK / KM RODADO.\nAba esperada: {ABA_MAXTRACK}",
    )
    arq_ativo = selecionar_arquivo_gui(
        "3 de 4 - Selecione a BASE DE ATIVOS / VEICULOS",
        f"3 de 4\n\nSelecione a BASE DE ATIVOS / VEICULOS.\nAba esperada: {ABA_ATIVO}",
    )
    arq_producao = selecionar_arquivo_gui(
        "4 de 4 - Selecione a BASE CLIENTE / PRODUCAO",
        f"4 de 4\n\nSelecione a BASE CLIENTE / PRODUCAO OFICIAL.\nAba esperada: {ABA_PRODUCAO}",
    )
    pasta_saida = os.path.dirname(arq_combustivel) or os.getcwd()
    arq_saida = selecionar_saida_gui(os.path.join(pasta_saida, "Resumo_Abastecimento_Odometro_V15.xlsx"))
    print("\nArquivos selecionados:")
    print(f"Combustivel: {arq_combustivel}")
    print(f"Maxtrack:    {arq_maxtrack}")
    print(f"Ativo:       {arq_ativo}")
    print(f"Cliente:     {arq_producao}")
    print(f"Saida:       {arq_saida}")
    return arq_combustivel, arq_maxtrack, arq_ativo, arq_producao, arq_saida


def normalizar_nome_coluna(valor):
    if valor is None:
        return ""
    txt = str(valor).strip().upper()
    txt = txt.replace("\n", " ").replace("\r", " ").replace("\t", " ")
    txt = " ".join(txt.split())
    return txt.replace(" ", "").replace("-", "").replace("_", "")


def normalizar_placa(valor):
    if pd.isna(valor):
        return None
    txt = str(valor).strip().upper().replace("-", "").replace(" ", "")
    return txt if txt else None


def extrair_go_do_veiculo(valor):
    if pd.isna(valor):
        return None
    txt = "".join(ch for ch in str(valor) if ch.isdigit())
    return txt[-5:] if txt else None


def para_numero(valor):
    if pd.isna(valor):
        return np.nan
    if isinstance(valor, (int, float, np.integer, np.floating)):
        return float(valor)
    txt = str(valor).strip()
    try:
        return float(txt.replace(".", "").replace(",", "."))
    except Exception:
        try:
            return float(txt)
        except Exception:
            return np.nan


def escolher_coluna(df, candidatos, obrigatoria=True):
    cols_original = list(df.columns)
    cols_norm = {normalizar_nome_coluna(c): c for c in cols_original}
    for candidato in candidatos:
        cand_norm = normalizar_nome_coluna(candidato)
        if cand_norm in cols_norm:
            return cols_norm[cand_norm]
        for col_norm, col_real in cols_norm.items():
            if cand_norm in col_norm or col_norm in cand_norm:
                return col_real
    if obrigatoria:
        print("\nColunas disponiveis:")
        for c in cols_original:
            print(f" - {c}")
        raise KeyError(f"Nenhuma das colunas foi encontrada: {candidatos}")
    return None


def escolher_melhor_coluna_por_palavras(df, lista_de_palavras, obrigatoria=True):
    palavras_alvo = set(p.upper() for p in lista_de_palavras)
    melhor_coluna = None
    melhor_score = -1
    for col in df.columns:
        base = str(col).upper().replace("\n", " ").replace("\r", " ").replace("\t", " ")
        base = base.replace("-", " ").replace("_", " ").replace("/", " ")
        score = len(palavras_alvo.intersection(set(base.split())))
        if score > melhor_score:
            melhor_score = score
            melhor_coluna = col
    if melhor_score > 0:
        return melhor_coluna
    if obrigatoria:
        raise KeyError(f"Nenhuma coluna compativel com palavras: {lista_de_palavras}")
    return None


def ler_excel_com_aba(caminho, aba_preferida, usecols=None):
    try:
        return pd.read_excel(caminho, sheet_name=aba_preferida, usecols=usecols)
    except Exception:
        xls = pd.ExcelFile(caminho)
        abas = xls.sheet_names
        print(f"\nNao foi possivel ler a aba '{aba_preferida}'.")
        for idx, aba in enumerate(abas, start=1):
            print(f"{idx} - {aba}")
        if len(abas) == 1:
            return pd.read_excel(caminho, sheet_name=abas[0], usecols=usecols)
        escolha = input("\nDigite o numero da aba correta: ").strip()
        return pd.read_excel(caminho, sheet_name=abas[int(escolha) - 1], usecols=usecols)


def excel_serial_para_datetime(valor):
    if valor is None or valor == "":
        return pd.NaT
    if isinstance(valor, (pd.Timestamp, datetime)):
        return pd.to_datetime(valor, errors="coerce")
    try:
        return pd.Timestamp(datetime(1899, 12, 30) + timedelta(days=float(valor)))
    except Exception:
        return pd.to_datetime(valor, errors="coerce")


def _xlsx_coluna_ref(ref):
    return re.sub(r"\d+", "", str(ref))


def _carregar_shared_strings_xlsx(zip_obj):
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    strings = []
    try:
        with zip_obj.open("xl/sharedStrings.xml") as f:
            for _, elem in ET.iterparse(f, events=("end",)):
                if elem.tag == ns + "si":
                    partes = []
                    for tnode in elem.iter(ns + "t"):
                        if tnode.text:
                            partes.append(tnode.text)
                    strings.append("".join(partes))
                    elem.clear()
    except KeyError:
        return []
    return strings


def ler_maxtrack_rapido_xlsx(caminho):
    cache_csv = os.path.splitext(caminho)[0] + ".cache.csv"
    if os.path.exists(cache_csv):
        print(f"Usando cache Maxtrack: {cache_csv}")
        df_cache = pd.read_csv(cache_csv)
        if "Início" in df_cache.columns:
            df_cache["Início"] = df_cache["Início"].apply(excel_serial_para_datetime)
        if "Fim" in df_cache.columns:
            df_cache["Fim"] = df_cache["Fim"].apply(excel_serial_para_datetime)
        for c in ["Odômetro Inicial", "Odômetro Final"]:
            if c in df_cache.columns:
                df_cache[c] = df_cache[c].apply(para_numero)
        return df_cache

    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    colunas = {
        "F": "Identificador/Placa",
        "I": "Início",
        "L": "Fim",
        "AF": "Odômetro Inicial",
        "AG": "Odômetro Final",
    }
    registros = []
    with zipfile.ZipFile(caminho) as z:
        shared = _carregar_shared_strings_xlsx(z)
        with z.open("xl/worksheets/sheet1.xml") as f:
            total = 0
            for _, elem in ET.iterparse(f, events=("end",)):
                if elem.tag == ns + "row":
                    numero_linha = int(elem.attrib.get("r", "0"))
                    if numero_linha > 1:
                        item = {nome: None for nome in colunas.values()}
                        for celula in elem.findall(ns + "c"):
                            col = _xlsx_coluna_ref(celula.attrib.get("r", ""))
                            if col not in colunas:
                                continue
                            v = celula.find(ns + "v")
                            valor = v.text if v is not None else None
                            if celula.attrib.get("t") == "s" and valor is not None:
                                try:
                                    valor = shared[int(valor)]
                                except Exception:
                                    pass
                            item[colunas[col]] = valor
                        registros.append(item)
                        total += 1
                        if total % 50000 == 0:
                            mostrar_progresso("Leitura rapida Maxtrack", total, 153762)
                    elem.clear()
    mostrar_progresso("Leitura rapida Maxtrack", len(registros), len(registros))
    df = pd.DataFrame(registros)
    df["Início"] = df["Início"].apply(excel_serial_para_datetime)
    df["Fim"] = df["Fim"].apply(excel_serial_para_datetime)
    df["Odômetro Inicial"] = df["Odômetro Inicial"].apply(para_numero)
    df["Odômetro Final"] = df["Odômetro Final"].apply(para_numero)
    return df


def periodo_interno(dt):
    if pd.isna(dt):
        return None
    return "AM" if pd.Timestamp(dt).hour < 12 else "PM"


def classificar_confianca_temporal(dif_min):
    if pd.isna(dif_min):
        return "NENHUM"
    if dif_min <= 10:
        return "ALTA"
    if dif_min <= 30:
        return "MEDIA"
    if dif_min <= 120:
        return "BAIXA"
    return "MUITO_BAIXA"


def preparar_bases(arq_combustivel, arq_maxtrack, arq_ativo, arq_producao):
    imprimir_etapa("Lendo e tratando bases")
    df_comb = ler_excel_com_aba(arq_combustivel, ABA_COMBUSTIVEL, usecols=None)
    df_ativo = ler_excel_com_aba(arq_ativo, ABA_ATIVO, usecols=None)
    df_prod = ler_excel_com_aba(arq_producao, ABA_PRODUCAO, usecols=None)
    df_max = ler_maxtrack_rapido_xlsx(arq_maxtrack)

    col_comb_veiculo = escolher_coluna(df_comb, ["Veiculo", "Veículo"])
    col_comb_data = escolher_coluna(df_comb, ["Data", "Data Ajustada", "Data Abastecimento"])
    col_comb_litros = escolher_coluna(df_comb, ["Volume (L)", "Litros", "Volume"])

    col_ativo_go = escolher_coluna(df_ativo, ["GO", "G.O", "CODIGO GO", "CODIGO"])
    try:
        col_ativo_placa = escolher_coluna(df_ativo, ["PLACA", "Placa", "Placa do Veiculo", "Placa do Veículo"])
    except KeyError:
        col_ativo_placa = escolher_melhor_coluna_por_palavras(df_ativo, ["PLACA"])

    col_prod_placa = escolher_coluna(df_prod, ["Placa"])
    col_prod_saida = escolher_coluna(df_prod, ["Saída", "Saida", "Hora Saída", "Hora Saida", "Data Saida"])
    col_prod_dist = escolher_coluna(df_prod, ["Distância", "Distancia", "Distancia KM", "KM"])

    ativo = df_ativo[[col_ativo_go, col_ativo_placa]].copy()
    ativo.columns = ["GO", "PLACA"]
    ativo["GO"] = ativo["GO"].apply(lambda x: "".join(ch for ch in str(x) if ch.isdigit()) if pd.notna(x) else None)
    ativo["PLACA"] = ativo["PLACA"].apply(normalizar_placa)
    ativo = ativo.dropna(subset=["GO", "PLACA"]).drop_duplicates()

    comb = df_comb[[col_comb_veiculo, col_comb_data, col_comb_litros]].copy()
    comb.columns = ["VEICULO", "DATA_HORA_ABASTECIMENTO", "LITROS"]
    comb["GO"] = comb["VEICULO"].apply(extrair_go_do_veiculo)
    comb["DATA_HORA_ABASTECIMENTO"] = pd.to_datetime(comb["DATA_HORA_ABASTECIMENTO"], errors="coerce")
    comb["LITROS"] = comb["LITROS"].apply(para_numero)
    comb = comb.merge(ativo, on="GO", how="left")
    comb["PLACA"] = comb["PLACA"].apply(normalizar_placa).fillna("SEM_PLACA")
    comb = comb.dropna(subset=["DATA_HORA_ABASTECIMENTO"]).copy()
    comb = comb.sort_values(["PLACA", "DATA_HORA_ABASTECIMENTO"], na_position="last").reset_index(drop=True)
    comb["ID_LINHA_ORIGINAL"] = comb.index.astype(str)

    eventos = []
    tmp_ini = df_max[["Identificador/Placa", "Início", "Odômetro Inicial"]].copy()
    tmp_ini.columns = ["PLACA", "DATA_HORA_REF", "ODOMETRO"]
    tmp_ini["TIPO_REF"] = "INICIO"
    eventos.append(tmp_ini)
    tmp_fim = df_max[["Identificador/Placa", "Fim", "Odômetro Final"]].copy()
    tmp_fim.columns = ["PLACA", "DATA_HORA_REF", "ODOMETRO"]
    tmp_fim["TIPO_REF"] = "FIM"
    eventos.append(tmp_fim)
    odom = pd.concat(eventos, ignore_index=True)
    odom["PLACA"] = odom["PLACA"].apply(normalizar_placa)
    odom["DATA_HORA_REF"] = pd.to_datetime(odom["DATA_HORA_REF"], errors="coerce")
    odom["ODOMETRO"] = odom["ODOMETRO"].apply(para_numero)
    odom = odom.dropna(subset=["PLACA", "DATA_HORA_REF", "ODOMETRO"]).copy()
    odom = odom[odom["ODOMETRO"] > 0].copy()
    odom = odom.drop_duplicates(subset=["PLACA", "DATA_HORA_REF", "ODOMETRO", "TIPO_REF"])
    odom = odom.sort_values(["PLACA", "DATA_HORA_REF"]).reset_index(drop=True)

    prod = df_prod[[col_prod_placa, col_prod_saida, col_prod_dist]].copy()
    prod.columns = ["PLACA", "DATA_HORA_PROD", "DISTANCIA_PROD"]
    prod["PLACA"] = prod["PLACA"].apply(normalizar_placa)
    prod["DATA_HORA_PROD"] = pd.to_datetime(prod["DATA_HORA_PROD"], errors="coerce")
    prod["DISTANCIA_PROD"] = prod["DISTANCIA_PROD"].apply(para_numero)
    prod = prod.dropna(subset=["PLACA", "DATA_HORA_PROD", "DISTANCIA_PROD"]).copy()
    prod = prod[prod["DISTANCIA_PROD"] >= 0].copy()
    prod = prod.sort_values(["PLACA", "DATA_HORA_PROD"]).reset_index(drop=True)
    prod["DIA_PROD_INTERNO"] = prod["DATA_HORA_PROD"].dt.date
    prod["TOTAL_DIA_PLACA_INTERNO"] = prod.groupby(["PLACA", "DIA_PROD_INTERNO"])["DISTANCIA_PROD"].transform("sum")

    print(f"Combustivel tratado: {len(comb)} registros")
    print(f"Maxtrack tratado: {len(odom)} eventos")
    print(f"Cliente/Producao tratada: {len(prod)} registros")
    return comb, odom, prod


def detalhar_distancia_producao(df_prod_placa, dt_ini, dt_fim):
    base = {
        "KM_CLIENTE_BRUTO": np.nan,
        "KM_CLIENTE_AJUSTADO": np.nan,
        "FLAG_CLIENTE_DIA_MAIOR_700": "NAO",
        "MOTIVO_CLIENTE_INTERVALO": None,
    }
    if df_prod_placa is None or df_prod_placa.empty:
        return base
    if pd.isna(dt_ini) or pd.isna(dt_fim) or dt_fim < dt_ini:
        return base
    janela = df_prod_placa[(df_prod_placa["DATA_HORA_PROD"] > dt_ini) & (df_prod_placa["DATA_HORA_PROD"] <= dt_fim)].copy()
    if janela.empty:
        return base

    km_bruto = float(janela["DISTANCIA_PROD"].sum())
    km_ajustado = 0.0
    flag = "NAO"
    motivos = []

    for dia, bloco in janela.groupby("DIA_PROD_INTERNO"):
        contrib = float(bloco["DISTANCIA_PROD"].sum())
        total_dia_vals = df_prod_placa.loc[df_prod_placa["DIA_PROD_INTERNO"] == dia, "TOTAL_DIA_PLACA_INTERNO"]
        total_dia = float(total_dia_vals.iloc[0]) if len(total_dia_vals) else contrib
        if total_dia > LIMITE_CLIENTE_DIA_ALTO:
            flag = "SIM"
            qtd_blocos = max(1, int(np.ceil(total_dia / LIMITE_CLIENTE_DIA_ALTO)))
            teto_bloco = float(np.ceil(total_dia / qtd_blocos))
            inicio_dia = pd.Timestamp(dia)
            fim_dia = inicio_dia + pd.Timedelta(days=1)
            ini_efetivo = max(pd.Timestamp(dt_ini), inicio_dia)
            fim_efetivo = min(pd.Timestamp(dt_fim), fim_dia)
            horas = max(0.01, (fim_efetivo - ini_efetivo).total_seconds() / 3600.0)
            blocos_intervalo = max(1, int(np.ceil(horas / 12.0)))
            teto_intervalo = min(MAX_SALTO_ABAST, teto_bloco * blocos_intervalo)
            contrib_ajustada = min(contrib, teto_intervalo)
            if contrib_ajustada < contrib:
                motivos.append(f"CLIENTE_DIA>{LIMITE_CLIENTE_DIA_ALTO}_CHAVEADO_{dia}:{round(contrib,2)}->{round(contrib_ajustada,2)}")
            else:
                motivos.append(f"CLIENTE_DIA>{LIMITE_CLIENTE_DIA_ALTO}_SEM_CORTE_{dia}")
            km_ajustado += contrib_ajustada
        else:
            km_ajustado += contrib

    if km_ajustado > MAX_SALTO_ABAST:
        motivos.append(f"CLIENTE_INTERVALO>{MAX_SALTO_ABAST}_LIMITADO:{round(km_ajustado, 2)}->{MAX_SALTO_ABAST}")
        km_ajustado = float(MAX_SALTO_ABAST)

    base["KM_CLIENTE_BRUTO"] = km_bruto
    base["KM_CLIENTE_AJUSTADO"] = float(km_ajustado)
    base["FLAG_CLIENTE_DIA_MAIOR_700"] = flag
    base["MOTIVO_CLIENTE_INTERVALO"] = "|".join(motivos) if motivos else None
    return base


def mediana_km_validos(grupo, idx_atual):
    if "KM_ENTRE_ABAST" not in grupo.columns:
        return np.nan
    inicio = max(0, idx_atual - JANELA_TENDENCIA)
    vals = grupo.loc[inicio:idx_atual - 1, "KM_ENTRE_ABAST"].dropna().tolist()
    vals = [float(v) for v in vals if 1 < float(v) <= MAX_SALTO_ABAST]
    return float(np.median(vals)) if vals else np.nan


def preparar_indices(odom, prod):
    eventos = {}
    for placa, g in odom.groupby("PLACA"):
        g = g.sort_values("DATA_HORA_REF").reset_index(drop=True)
        eventos[placa] = {
            "df": g,
            "ts": g["DATA_HORA_REF"].astype("int64").to_numpy(),
        }
    prod_map = {placa: g.sort_values("DATA_HORA_PROD").reset_index(drop=True) for placa, g in prod.groupby("PLACA")}
    return eventos, prod_map


def listar_candidatos_match_rapido(idx_eventos, placa, dt_abast):
    if not placa or placa not in idx_eventos or pd.isna(dt_abast):
        return pd.DataFrame()
    item = idx_eventos[placa]
    g = item["df"]
    ts = item["ts"]
    centro = pd.Timestamp(dt_abast).value
    janela_ns = int(pd.Timedelta(minutes=JANELA_MATCH_MINUTOS).value)
    ini = np.searchsorted(ts, centro - janela_ns, side="left")
    fim = np.searchsorted(ts, centro + janela_ns, side="right")
    if fim <= ini:
        return pd.DataFrame()
    cand = g.iloc[ini:fim].copy()
    cand["DIFERENCA_MINUTOS"] = (cand["DATA_HORA_REF"] - dt_abast).abs().dt.total_seconds() / 60.0
    cand["_MESMO_DIA_INTERNO"] = cand["DATA_HORA_REF"].dt.date == pd.Timestamp(dt_abast).date()
    periodo_abast = periodo_interno(dt_abast)
    cand["_MESMO_PERIODO_INTERNO"] = cand["DATA_HORA_REF"].apply(lambda x: periodo_interno(x) == periodo_abast)
    cand["_SCORE_INICIAL"] = 0.0
    cand["_SCORE_INICIAL"] += (1 - (cand["DIFERENCA_MINUTOS"].clip(0, JANELA_MATCH_MINUTOS) / JANELA_MATCH_MINUTOS)) * PESO_TEMPO
    cand["_SCORE_INICIAL"] += np.where(cand["_MESMO_DIA_INTERNO"], PESO_MESMO_DIA, 0)
    cand["_SCORE_INICIAL"] += np.where(cand["_MESMO_PERIODO_INTERNO"], PESO_MESMO_PERIODO, 0)
    cand["_SCORE_INICIAL"] += np.where(cand["TIPO_REF"] == "FIM", PESO_TIPO_REF, PESO_TIPO_REF * 0.75)
    return cand.sort_values(["_SCORE_INICIAL", "DIFERENCA_MINUTOS", "DATA_HORA_REF"], ascending=[False, True, False]).head(TOP_CANDIDATOS_POR_ABAST).reset_index(drop=True)


def buscar_matches(comb, idx_eventos):
    imprimir_etapa("Buscando candidatos Maxtrack por abastecimento")
    partes = []
    grupos = list(comb.groupby("PLACA", dropna=False))
    total_grupos = len(grupos)

    for pos, (placa, grupo) in enumerate(grupos, start=1):
        g = grupo.sort_values("DATA_HORA_ABASTECIMENTO").copy().reset_index(drop=True)

        if (not placa) or placa == "SEM_PLACA" or placa not in idx_eventos:
            g["ODOMETRO_CANDIDATO"] = np.nan
            g["DATA_HORA_MAXTRACK_ENCONTRADA"] = pd.NaT
            g["TIPO_REF_MAXTRACK"] = None
            g["DIFERENCA_MINUTOS"] = np.nan
            g["STATUS_MATCH"] = "SEM_PLACA" if (not placa or placa == "SEM_PLACA") else "NAO_ENCONTRADO"
            g["CONFIANCA_TEMPORAL"] = "NENHUM"
            g["SCORE_MATCH"] = np.nan
            partes.append(g)
            continue

        eventos = idx_eventos[placa]["df"][["DATA_HORA_REF", "ODOMETRO", "TIPO_REF"]].sort_values("DATA_HORA_REF").copy()
        m = pd.merge_asof(
            g,
            eventos,
            left_on="DATA_HORA_ABASTECIMENTO",
            right_on="DATA_HORA_REF",
            direction="nearest",
            tolerance=pd.Timedelta(minutes=JANELA_MATCH_MINUTOS),
        )

        m["ODOMETRO_CANDIDATO"] = m["ODOMETRO"]
        m["DATA_HORA_MAXTRACK_ENCONTRADA"] = m["DATA_HORA_REF"]
        m["TIPO_REF_MAXTRACK"] = m["TIPO_REF"]
        m["DIFERENCA_MINUTOS"] = (
            (m["DATA_HORA_MAXTRACK_ENCONTRADA"] - m["DATA_HORA_ABASTECIMENTO"]).abs().dt.total_seconds() / 60.0
        )
        m["STATUS_MATCH"] = np.where(m["ODOMETRO_CANDIDATO"].notna(), "ENCONTRADO", "NAO_ENCONTRADO")
        m["CONFIANCA_TEMPORAL"] = m["DIFERENCA_MINUTOS"].apply(classificar_confianca_temporal)

        mesmo_dia = m["DATA_HORA_MAXTRACK_ENCONTRADA"].dt.date == m["DATA_HORA_ABASTECIMENTO"].dt.date
        mesmo_periodo = [
            periodo_interno(a) == periodo_interno(b)
            for a, b in zip(m["DATA_HORA_ABASTECIMENTO"], m["DATA_HORA_MAXTRACK_ENCONTRADA"])
        ]
        score = (1 - (m["DIFERENCA_MINUTOS"].fillna(JANELA_MATCH_MINUTOS).clip(0, JANELA_MATCH_MINUTOS) / JANELA_MATCH_MINUTOS)) * PESO_TEMPO
        score += np.where(mesmo_dia, PESO_MESMO_DIA, 0)
        score += np.where(mesmo_periodo, PESO_MESMO_PERIODO, 0)
        score += np.where(m["TIPO_REF_MAXTRACK"] == "FIM", PESO_TIPO_REF, PESO_TIPO_REF * 0.75)
        m["SCORE_MATCH"] = np.where(m["ODOMETRO_CANDIDATO"].notna(), score, np.nan)

        m = m.drop(columns=["ODOMETRO", "DATA_HORA_REF", "TIPO_REF"], errors="ignore")
        partes.append(m)

        if pos % 10 == 0 or pos == total_grupos:
            mostrar_progresso("Match Maxtrack por placa", pos, total_grupos)

    resultado = pd.concat(partes, ignore_index=True)
    return resultado.sort_values(["PLACA", "DATA_HORA_ABASTECIMENTO"], na_position="last").reset_index(drop=True)

def agrupar_abastecimentos_proximos(grupo):
    grupo = grupo.sort_values("DATA_HORA_ABASTECIMENTO").copy().reset_index(drop=True)
    if grupo.empty:
        return grupo
    blocos = []
    atual = [0]
    for i in range(1, len(grupo)):
        ref = grupo.loc[atual[-1]]
        cur = grupo.loc[i]
        delta_min = (cur["DATA_HORA_ABASTECIMENTO"] - ref["DATA_HORA_ABASTECIMENTO"]).total_seconds() / 60.0
        od_ref = ref["ODOMETRO_CANDIDATO"]
        od_cur = cur["ODOMETRO_CANDIDATO"]
        pode = False
        if delta_min <= MINUTOS_AGRUPAMENTO:
            if pd.notna(od_ref) and pd.notna(od_cur):
                pode = abs(float(od_cur) - float(od_ref)) <= MAX_DIFF_ODOM_AGRUPAMENTO
            else:
                pode = True
        if pode:
            atual.append(i)
        else:
            blocos.append(atual)
            atual = [i]
    blocos.append(atual)
    linhas = []
    for gid, idxs in enumerate(blocos, start=1):
        b = grupo.loc[idxs].copy()
        if b["ODOMETRO_CANDIDATO"].notna().any():
            melhor_idx = b.sort_values(["SCORE_MATCH", "DIFERENCA_MINUTOS"], ascending=[False, True]).index[0]
            melhor = b.loc[melhor_idx]
            od_cand = melhor["ODOMETRO_CANDIDATO"]
            dt_mt = melhor["DATA_HORA_MAXTRACK_ENCONTRADA"]
            tipo = melhor["TIPO_REF_MAXTRACK"]
            dif = melhor["DIFERENCA_MINUTOS"]
            status = "ENCONTRADO"
            conf = melhor["CONFIANCA_TEMPORAL"]
            score = melhor["SCORE_MATCH"]
        else:
            od_cand = np.nan
            dt_mt = pd.NaT
            tipo = None
            dif = np.nan
            status = b["STATUS_MATCH"].iloc[0]
            conf = "NENHUM"
            score = np.nan
        linhas.append({
            "PLACA": b["PLACA"].iloc[0],
            "GO": b["GO"].iloc[0],
            "VEICULO": b["VEICULO"].iloc[0],
            "DATA_HORA_ABASTECIMENTO": b["DATA_HORA_ABASTECIMENTO"].min(),
            "LITROS": b["LITROS"].sum(),
            "ODOMETRO_MATCH": np.nan,
            "ODOMETRO_CANDIDATO": od_cand,
            "DATA_HORA_MAXTRACK_ENCONTRADA": dt_mt,
            "TIPO_REF_MAXTRACK": tipo,
            "DIFERENCA_MINUTOS": dif,
            "STATUS_MATCH": status,
            "CONFIANCA_TEMPORAL": conf,
            "SCORE_MATCH": score,
            "QTD_EVENTOS_GRUPO": len(b),
            "AGRUPADO_DUPLICADO": "SIM" if len(b) > 1 else "NAO",
            "GRUPO_ABASTECIMENTO": gid,
            "IDS_LINHAS_ORIGEM": ",".join(str(x) for x in b["ID_LINHA_ORIGINAL"].tolist()),
        })
    return pd.DataFrame(linhas)


def consolidar_abastecimentos(resultado_match):
    imprimir_etapa("Consolidando abastecimentos proximos")
    partes = []
    grupos = list(resultado_match.groupby("PLACA", dropna=False))
    for pos, (_, grupo) in enumerate(grupos, start=1):
        partes.append(agrupar_abastecimentos_proximos(grupo))
        if pos % 10 == 0 or pos == len(grupos):
            mostrar_progresso("Consolidacao por placa", pos, len(grupos))
    return pd.concat(partes, ignore_index=True).sort_values(["PLACA", "DATA_HORA_ABASTECIMENTO"], na_position="last").reset_index(drop=True)


def preencher_grupo(grupo, df_prod_placa):
    g = grupo.sort_values("DATA_HORA_ABASTECIMENTO").copy().reset_index(drop=True)
    g["MATCH_VALIDADO"] = False
    g["EH_ANCORA"] = False
    g["ORIGEM_ODOMETRO_MATCH"] = None
    g["CONFIANCA_ODOMETRO_MATCH"] = "SEM_ODOMETRO"
    g["MOTIVO_ODOMETRO_MATCH"] = None

    for i in range(len(g)):
        if pd.notna(g.loc[i, "ODOMETRO_CANDIDATO"]):
            g.loc[i, "ODOMETRO_MATCH"] = round(float(g.loc[i, "ODOMETRO_CANDIDATO"]), 0)
            g.loc[i, "MATCH_VALIDADO"] = True
            g.loc[i, "EH_ANCORA"] = True
            g.loc[i, "ORIGEM_ODOMETRO_MATCH"] = "MAXTRACK_DIRETO"
            g.loc[i, "CONFIANCA_ODOMETRO_MATCH"] = "ALTA"
            g.loc[i, "MOTIVO_ODOMETRO_MATCH"] = "ANCORA_MAXTRACK"

    anchors = g.index[g["EH_ANCORA"] == True].tolist()
    if not anchors:
        g["ORIGEM_ODOMETRO_MATCH"] = "SEM_MAXTRACK"
        g["CONFIANCA_ODOMETRO_MATCH"] = "MUITO_BAIXA"
        g["MOTIVO_ODOMETRO_MATCH"] = "SEM_ANCORA_MAXTRACK"
        return g

    # Antes da primeira ancora
    first = anchors[0]
    od_first = g.loc[first, "ODOMETRO_MATCH"]
    dt_first = g.loc[first, "DATA_HORA_ABASTECIMENTO"]
    for i in range(first - 1, -1, -1):
        dt_i = g.loc[i, "DATA_HORA_ABASTECIMENTO"]
        det = detalhar_distancia_producao(df_prod_placa, dt_i, dt_first)
        dist = det["KM_CLIENTE_AJUSTADO"]
        passo = max(np.ceil(dist), MIN_INCREMENTO_ABSOLUTO) if pd.notna(dist) else MIN_INCREMENTO_SEM_CLIENTE * (first - i)
        passo = min(passo, MAX_SALTO_ABAST * (first - i))
        g.loc[i, "ODOMETRO_MATCH"] = round(float(od_first) - passo, 0)
        g.loc[i, "ORIGEM_ODOMETRO_MATCH"] = "CLIENTE_ATE_PRIMEIRO_MAXTRACK" if pd.notna(dist) else "TENDENCIA_FALLBACK"
        g.loc[i, "CONFIANCA_ODOMETRO_MATCH"] = "BAIXA" if pd.notna(dist) else "MUITO_BAIXA"
        g.loc[i, "MOTIVO_ODOMETRO_MATCH"] = "PREENCHIDO_ANTES_PRIMEIRA_ANCORA"

    # Entre ancoras
    for k in range(len(anchors) - 1):
        a, b = anchors[k], anchors[k + 1]
        od_a = float(g.loc[a, "ODOMETRO_MATCH"])
        od_b = float(g.loc[b, "ODOMETRO_MATCH"])
        dt_a = g.loc[a, "DATA_HORA_ABASTECIMENTO"]
        dt_b = g.loc[b, "DATA_HORA_ABASTECIMENTO"]
        delta_total = od_b - od_a
        if b - a <= 1:
            continue
        if delta_total <= 0:
            # Se a ancora futura parece regressiva, preenche por cliente/fallback a partir da anterior.
            for j in range(a + 1, b):
                det = detalhar_distancia_producao(df_prod_placa, g.loc[j - 1, "DATA_HORA_ABASTECIMENTO"], g.loc[j, "DATA_HORA_ABASTECIMENTO"])
                inc = max(np.ceil(det["KM_CLIENTE_AJUSTADO"]), MIN_INCREMENTO_ABSOLUTO) if pd.notna(det["KM_CLIENTE_AJUSTADO"]) else MIN_INCREMENTO_SEM_CLIENTE
                inc = min(inc, MAX_SALTO_ABAST)
                g.loc[j, "ODOMETRO_MATCH"] = round(float(g.loc[j - 1, "ODOMETRO_MATCH"]) + inc, 0)
                g.loc[j, "ORIGEM_ODOMETRO_MATCH"] = "CLIENTE_ENTRE_MAXTRACK_REGRESSIVO"
                g.loc[j, "CONFIANCA_ODOMETRO_MATCH"] = "BAIXA"
                g.loc[j, "MOTIVO_ODOMETRO_MATCH"] = "ANCORA_POSTERIOR_REGRESSIVA"
            continue
        det_total = detalhar_distancia_producao(df_prod_placa, dt_a, dt_b)
        dist_total = det_total["KM_CLIENTE_AJUSTADO"]
        usar_tempo = pd.isna(dist_total) or dist_total <= 0
        for j in range(a + 1, b):
            dt_j = g.loc[j, "DATA_HORA_ABASTECIMENTO"]
            if usar_tempo:
                dur_total = (dt_b - dt_a).total_seconds()
                dur_parcial = (dt_j - dt_a).total_seconds()
                prop = max(0, min(1, dur_parcial / dur_total)) if dur_total > 0 else 0
                od_calc = od_a + delta_total * prop
                origem = "TEMPO_ENTRE_MAXTRACK"
                motivo = "PREENCHIDO_POR_TEMPO_ENTRE_ANCORAS"
            else:
                dist_parcial = detalhar_distancia_producao(df_prod_placa, dt_a, dt_j)["KM_CLIENTE_AJUSTADO"]
                dist_restante = detalhar_distancia_producao(df_prod_placa, dt_j, dt_b)["KM_CLIENTE_AJUSTADO"]
                prop = max(0, min(1, dist_parcial / dist_total)) if pd.notna(dist_parcial) and dist_total > 0 else 0
                estimado = od_a + delta_total * prop
                minimo = od_a + max(np.ceil(dist_parcial), 0) if pd.notna(dist_parcial) else od_a
                maximo = od_b - max(np.ceil(dist_restante), 0) if pd.notna(dist_restante) else od_b
                od_calc = min(max(estimado, minimo), maximo) if minimo <= maximo else estimado
                origem = "CLIENTE_ENTRE_MAXTRACK"
                motivo = "PREENCHIDO_COM_CLIENTE_CHAVEADO_ENTRE_ANCORAS"
            g.loc[j, "ODOMETRO_MATCH"] = round(od_calc, 0)
            g.loc[j, "ORIGEM_ODOMETRO_MATCH"] = origem
            g.loc[j, "CONFIANCA_ODOMETRO_MATCH"] = "MEDIA"
            g.loc[j, "MOTIVO_ODOMETRO_MATCH"] = motivo

    # Depois da ultima ancora
    last = anchors[-1]
    for i in range(last + 1, len(g)):
        det = detalhar_distancia_producao(df_prod_placa, g.loc[i - 1, "DATA_HORA_ABASTECIMENTO"], g.loc[i, "DATA_HORA_ABASTECIMENTO"])
        dist = det["KM_CLIENTE_AJUSTADO"]
        inc = max(np.ceil(dist), MIN_INCREMENTO_ABSOLUTO) if pd.notna(dist) else MIN_INCREMENTO_SEM_CLIENTE
        inc = min(inc, MAX_SALTO_ABAST)
        g.loc[i, "ODOMETRO_MATCH"] = round(float(g.loc[i - 1, "ODOMETRO_MATCH"]) + inc, 0)
        g.loc[i, "ORIGEM_ODOMETRO_MATCH"] = "CLIENTE_APOS_ULTIMO_MAXTRACK" if pd.notna(dist) else "TENDENCIA_FALLBACK"
        g.loc[i, "CONFIANCA_ODOMETRO_MATCH"] = "BAIXA" if pd.notna(dist) else "MUITO_BAIXA"
        g.loc[i, "MOTIVO_ODOMETRO_MATCH"] = "PREENCHIDO_APOS_ULTIMA_ANCORA"
    return g


def ajustar_odometro_match_final(grupo, df_prod_placa):
    g = grupo.sort_values("DATA_HORA_ABASTECIMENTO").copy().reset_index(drop=True)
    g["KM_CLIENTE_INTERVALO"] = np.nan
    g["KM_CLIENTE_BRUTO_INTERVALO"] = np.nan
    g["FLAG_CLIENTE_DIA_MAIOR_700"] = "NAO"
    g["MOTIVO_CLIENTE_INTERVALO"] = None
    g["KM_ENTRE_ABAST"] = np.nan
    g["DIFERENCA_ODOM_X_CLIENTE"] = np.nan
    g["FLAG_INTERVALO_FORA_FAIXA"] = "NAO"
    g["AJUSTE_ODOMETRO_MATCH"] = "NAO"

    for i in range(1, len(g)):
        od_prev = g.loc[i - 1, "ODOMETRO_MATCH"]
        od_cur = g.loc[i, "ODOMETRO_MATCH"]
        if pd.isna(od_prev) or pd.isna(od_cur):
            continue
        dt_prev = g.loc[i - 1, "DATA_HORA_ABASTECIMENTO"]
        dt_cur = g.loc[i, "DATA_HORA_ABASTECIMENTO"]
        det = detalhar_distancia_producao(df_prod_placa, dt_prev, dt_cur)
        km_cliente = det["KM_CLIENTE_AJUSTADO"]
        km_bruto = det["KM_CLIENTE_BRUTO"]
        g.loc[i, "KM_CLIENTE_INTERVALO"] = km_cliente
        g.loc[i, "KM_CLIENTE_BRUTO_INTERVALO"] = km_bruto
        g.loc[i, "FLAG_CLIENTE_DIA_MAIOR_700"] = det["FLAG_CLIENTE_DIA_MAIOR_700"]
        g.loc[i, "MOTIVO_CLIENTE_INTERVALO"] = det["MOTIVO_CLIENTE_INTERVALO"]

        if pd.notna(km_cliente):
            min_ok = float(np.ceil(km_cliente + MARGEM_TECNICA_CLIENTE))
            conflito = None
            if min_ok > MAX_SALTO_ABAST:
                min_ok = MAX_SALTO_ABAST
                conflito = "CONFLITO_CLIENTE_AJUSTADO_MAIOR_1500_MAXTRACK_PRIORIZADO"
        else:
            tendencia = mediana_km_validos(g, i)
            if pd.isna(tendencia):
                min_ok = MIN_INCREMENTO_SEM_CLIENTE
            else:
                min_ok = max(MIN_INCREMENTO_SEM_CLIENTE, np.floor(tendencia * 0.8))
            min_ok = min(min_ok, MAX_VARIACAO_LOCAL_SEM_CLIENTE)
            conflito = None

        max_ok = MAX_SALTO_ABAST
        km_odom = float(od_cur) - float(od_prev)
        ajuste = "NAO"
        novo = float(od_cur)

        if km_odom <= 1:
            novo = float(od_prev) + max(min_ok, MIN_INCREMENTO_ABSOLUTO)
            ajuste = "AJUSTADO_KM_MENOR_OU_IGUAL_1"
        elif km_odom < min_ok:
            novo = float(od_prev) + min_ok
            ajuste = "AJUSTADO_PARA_DISTANCIA_CLIENTE_CHAVEADA"
        elif km_odom > max_ok:
            novo = float(od_prev) + max_ok
            ajuste = "AJUSTADO_SALTO_MAIOR_1500"

        if ajuste != "NAO":
            g.loc[i, "ODOMETRO_MATCH"] = round(novo, 0)
            g.loc[i, "AJUSTE_ODOMETRO_MATCH"] = ajuste
            g.loc[i, "FLAG_INTERVALO_FORA_FAIXA"] = "SIM"
            motivo = g.loc[i, "MOTIVO_ODOMETRO_MATCH"]
            motivo = "" if pd.isna(motivo) or motivo is None else str(motivo)
            extras = [ajuste]
            if conflito:
                extras.append(conflito)
            g.loc[i, "MOTIVO_ODOMETRO_MATCH"] = "|".join([x for x in [motivo] + extras if x])
            if g.loc[i, "EH_ANCORA"] == True:
                g.loc[i, "ORIGEM_ODOMETRO_MATCH"] = "MAXTRACK_AJUSTADO_REGRA"
                g.loc[i, "CONFIANCA_ODOMETRO_MATCH"] = "MEDIA"
            else:
                g.loc[i, "CONFIANCA_ODOMETRO_MATCH"] = "BAIXA"
        elif conflito:
            motivo = g.loc[i, "MOTIVO_ODOMETRO_MATCH"]
            motivo = "" if pd.isna(motivo) or motivo is None else str(motivo)
            g.loc[i, "MOTIVO_ODOMETRO_MATCH"] = "|".join([x for x in [motivo, conflito] if x])
            g.loc[i, "FLAG_INTERVALO_FORA_FAIXA"] = "SIM"

        km_final = float(g.loc[i, "ODOMETRO_MATCH"]) - float(od_prev)
        g.loc[i, "KM_ENTRE_ABAST"] = km_final
        if pd.notna(km_cliente):
            g.loc[i, "DIFERENCA_ODOM_X_CLIENTE"] = km_final - km_cliente
            if km_final + 1e-9 < km_cliente or km_final > MAX_SALTO_ABAST:
                g.loc[i, "FLAG_INTERVALO_FORA_FAIXA"] = "SIM"
        else:
            if km_final <= 1 or km_final > MAX_SALTO_ABAST:
                g.loc[i, "FLAG_INTERVALO_FORA_FAIXA"] = "SIM"
    return g


def gerar_resultado(comb, odom, prod):
    idx_eventos, prod_map = preparar_indices(odom, prod)
    resultado_match = buscar_matches(comb, idx_eventos)
    consolidado = consolidar_abastecimentos(resultado_match)
    imprimir_etapa("Preenchendo ODOMETRO_MATCH e aplicando regras finais")
    partes = []
    grupos = list(consolidado.groupby("PLACA", dropna=False))
    for pos, (placa, grupo) in enumerate(grupos, start=1):
        df_prod_placa = prod_map.get(placa)
        preenchido = preencher_grupo(grupo, df_prod_placa)
        ajustado = ajustar_odometro_match_final(preenchido, df_prod_placa)
        partes.append(ajustado)
        if pos % 10 == 0 or pos == len(grupos):
            mostrar_progresso("Processamento por placa", pos, len(grupos))
    final = pd.concat(partes, ignore_index=True).sort_values(["PLACA", "DATA_HORA_ABASTECIMENTO"], na_position="last").reset_index(drop=True)
    return final, consolidado


def classificar_diff_mes(x):
    if pd.isna(x):
        return "SEM_COMPARACAO"
    x = abs(float(x))
    if x <= 100:
        return "OTIMA"
    if x <= 250:
        return "BOA"
    if x <= 500:
        return "ATENCAO"
    return "CRITICA"


def gerar_resumo_mensal(resultado_final, prod):
    base = resultado_final[resultado_final["ODOMETRO_MATCH"].notna()].copy()
    base["ANO_MES"] = base["DATA_HORA_ABASTECIMENTO"].dt.to_period("M").astype(str)
    resumo_odom = (
        base.sort_values(["PLACA", "DATA_HORA_ABASTECIMENTO"])
        .groupby(["PLACA", "ANO_MES"], as_index=False)
        .agg(
            PRIMEIRA_DATA_MES=("DATA_HORA_ABASTECIMENTO", "min"),
            ULTIMA_DATA_MES=("DATA_HORA_ABASTECIMENTO", "max"),
            PRIMEIRO_ODOMETRO_MES=("ODOMETRO_MATCH", "first"),
            ULTIMO_ODOMETRO_MES=("ODOMETRO_MATCH", "last"),
            QTD_ABAST_MES=("ODOMETRO_MATCH", "size"),
        )
    )
    resumo_odom["KM_RODADO_ODOMETRO_MES"] = resumo_odom["ULTIMO_ODOMETRO_MES"] - resumo_odom["PRIMEIRO_ODOMETRO_MES"]
    prod_mes = prod.copy()
    prod_mes["ANO_MES"] = prod_mes["DATA_HORA_PROD"].dt.to_period("M").astype(str)
    resumo_prod = (
        prod_mes.groupby(["PLACA", "ANO_MES"], as_index=False)
        .agg(
            DISTANCIA_CLIENTE_MES=("DISTANCIA_PROD", "sum"),
            PRIMEIRA_DATA_CLIENTE_MES=("DATA_HORA_PROD", "min"),
            ULTIMA_DATA_CLIENTE_MES=("DATA_HORA_PROD", "max"),
            QTD_EVENTOS_CLIENTE_MES=("DISTANCIA_PROD", "size"),
        )
    )
    resumo = resumo_odom.merge(resumo_prod, on=["PLACA", "ANO_MES"], how="outer")
    resumo["DIFERENCA_ODOM_X_CLIENTE_MES"] = resumo["KM_RODADO_ODOMETRO_MES"] - resumo["DISTANCIA_CLIENTE_MES"]
    resumo["DIFERENCA_ABS_MES"] = resumo["DIFERENCA_ODOM_X_CLIENTE_MES"].abs()
    resumo["CLASSIFICACAO_DIFERENCA_MES"] = resumo["DIFERENCA_ODOM_X_CLIENTE_MES"].apply(classificar_diff_mes)
    return resumo.sort_values(["PLACA", "ANO_MES"]).reset_index(drop=True)


def gerar_indicadores(resultado_final, comb):
    return pd.DataFrame({
        "Indicador": [
            "Total abastecimentos originais",
            "Total apos consolidacao",
            "Eventos agrupados",
            "ODOMETRO_MATCH preenchido",
            "ODOMETRO_MATCH vazio",
            "Matches Maxtrack validados",
            "Anchors Maxtrack",
            "Intervalos fora da faixa",
            "Saltos > 1500",
            "Intervalos abaixo da distancia cliente ajustada",
            "Ajustes no ODOMETRO_MATCH",
            "Intervalos com Cliente bruto dia > 700",
        ],
        "Valor": [
            len(comb),
            len(resultado_final),
            int((resultado_final["AGRUPADO_DUPLICADO"] == "SIM").sum()),
            int(resultado_final["ODOMETRO_MATCH"].notna().sum()),
            int(resultado_final["ODOMETRO_MATCH"].isna().sum()),
            int((resultado_final["MATCH_VALIDADO"] == True).sum()),
            int((resultado_final["EH_ANCORA"] == True).sum()),
            int((resultado_final["FLAG_INTERVALO_FORA_FAIXA"] == "SIM").sum()),
            int((resultado_final["KM_ENTRE_ABAST"] > MAX_SALTO_ABAST).sum()),
            int(((resultado_final["DIFERENCA_ODOM_X_CLIENTE"] < -1e-9) & resultado_final["DIFERENCA_ODOM_X_CLIENTE"].notna()).sum()),
            int((resultado_final["AJUSTE_ODOMETRO_MATCH"] != "NAO").sum()),
            int((resultado_final["FLAG_CLIENTE_DIA_MAIOR_700"] == "SIM").sum()),
        ],
    })


def formatar_excel(caminho):
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        wb = load_workbook(caminho)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 10
                for cell in ws[col_letter][:200]:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 42)
        wb.save(caminho)
    except Exception as e:
        print(f"Arquivo gerado, mas nao foi possivel aplicar formatacao: {e}")


def exportar(arq_saida, resultado_final, resumo_mensal, indicadores, comb, odom, prod):
    resumo_cols = [
        "PLACA", "DATA_HORA_ABASTECIMENTO", "LITROS", "ODOMETRO_MATCH", "KM_ENTRE_ABAST",
        "KM_CLIENTE_INTERVALO", "KM_CLIENTE_BRUTO_INTERVALO", "FLAG_CLIENTE_DIA_MAIOR_700",
        "MOTIVO_CLIENTE_INTERVALO", "DIFERENCA_ODOM_X_CLIENTE", "ORIGEM_ODOMETRO_MATCH",
        "CONFIANCA_ODOMETRO_MATCH", "MOTIVO_ODOMETRO_MATCH", "AJUSTE_ODOMETRO_MATCH",
        "FLAG_INTERVALO_FORA_FAIXA", "MATCH_VALIDADO", "EH_ANCORA", "STATUS_MATCH",
        "CONFIANCA_TEMPORAL", "AGRUPADO_DUPLICADO", "QTD_EVENTOS_GRUPO",
        "DATA_HORA_MAXTRACK_ENCONTRADA", "TIPO_REF_MAXTRACK", "DIFERENCA_MINUTOS",
        "SCORE_MATCH", "GO", "VEICULO", "GRUPO_ABASTECIMENTO", "IDS_LINHAS_ORIGEM",
    ]
    resumo = resultado_final[resumo_cols].copy()
    with pd.ExcelWriter(arq_saida, engine="openpyxl") as writer:
        resumo.to_excel(writer, sheet_name="Resumo", index=False)
        resultado_final.to_excel(writer, sheet_name="Base_Completa", index=False)
        resumo_mensal.to_excel(writer, sheet_name="Resumo_Mensal", index=False)
        indicadores.to_excel(writer, sheet_name="Indicadores", index=False)
        comb.to_excel(writer, sheet_name="Comb_Tratado", index=False)
        odom.to_excel(writer, sheet_name="Eventos_Maxtrack", index=False)
        prod.drop(columns=["DIA_PROD_INTERNO", "TOTAL_DIA_PLACA_INTERNO"], errors="ignore").to_excel(writer, sheet_name="Producao_Tratada", index=False)
    formatar_excel(arq_saida)


def main():
    arq_combustivel, arq_maxtrack, arq_ativo, arq_producao, arq_saida = selecionar_bases()
    comb, odom, prod = preparar_bases(arq_combustivel, arq_maxtrack, arq_ativo, arq_producao)
    resultado_final, _ = gerar_resultado(comb, odom, prod)
    imprimir_etapa("Gerando resumos e exportando")
    resumo_mensal = gerar_resumo_mensal(resultado_final, prod)
    indicadores = gerar_indicadores(resultado_final, comb)
    exportar(arq_saida, resultado_final, resumo_mensal, indicadores, comb, odom, prod)
    print("\nIndicadores finais:")
    print(indicadores.to_string(index=False))
    print("\n" + "=" * 70)
    print("Arquivo gerado com sucesso!")
    print(arq_saida)
    print("=" * 70)



def salvar_upload_temporario(uploaded_file, prefixo):
    suffix = Path(uploaded_file.name).suffix or ".xlsx"
    temp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix, prefix=prefixo)
    temp.write(uploaded_file.getbuffer())
    temp.flush()
    temp.close()
    return temp.name


def validar_uploads(arq_combustivel, arq_maxtrack, arq_ativo, arq_producao):
    faltantes = []
    if arq_combustivel is None:
        faltantes.append("1 de 4 - Base de dados - Combustivel.xlsx")
    if arq_maxtrack is None:
        faltantes.append("2 de 4 - Base de dados - Km Rodado Maxtrack.xlsx")
    if arq_ativo is None:
        faltantes.append("3 de 4 - Base de dados - Ativo de Veiculos.xlsx")
    if arq_producao is None:
        faltantes.append("4 de 4 - Producao Oficial - Relatorio Cliente.xlsx")
    return faltantes


def processar_streamlit(arq_combustivel_up, arq_maxtrack_up, arq_ativo_up, arq_producao_up, nome_saida):
    progress_bar = st.progress(0, text="0% - Aguardando inicio")
    status_box = st.empty()
    log_box = st.empty()
    configurar_progresso_streamlit(progress_bar, status_box, log_box)

    atualizar_progresso_streamlit(2, "Salvando arquivos enviados temporariamente")
    registrar_log_streamlit("Recebendo arquivos selecionados pelo usuario")

    arq_combustivel = salvar_upload_temporario(arq_combustivel_up, "combustivel_")
    arq_maxtrack = salvar_upload_temporario(arq_maxtrack_up, "maxtrack_")
    arq_ativo = salvar_upload_temporario(arq_ativo_up, "ativo_")
    arq_producao = salvar_upload_temporario(arq_producao_up, "producao_")

    saida_temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="resultado_odometro_")
    arq_saida = saida_temp.name
    saida_temp.close()

    atualizar_progresso_streamlit(4, "Arquivos recebidos. Iniciando leitura e tratamento das bases")

    comb, odom, prod = preparar_bases(arq_combustivel, arq_maxtrack, arq_ativo, arq_producao)
    registrar_log_streamlit(f"Combustivel tratado: {len(comb)} registros")
    registrar_log_streamlit(f"Maxtrack tratado: {len(odom)} eventos")
    registrar_log_streamlit(f"Cliente/Producao tratada: {len(prod)} registros")

    atualizar_progresso_streamlit(34, "Bases tratadas. Iniciando vinculo Combustivel x Maxtrack x Cliente")
    resultado_final, _ = gerar_resultado(comb, odom, prod)

    imprimir_etapa("Gerando resumos e exportando")
    resumo_mensal = gerar_resumo_mensal(resultado_final, prod)
    indicadores = gerar_indicadores(resultado_final, comb)

    atualizar_progresso_streamlit(92, "Exportando Excel final")
    exportar(arq_saida, resultado_final, resumo_mensal, indicadores, comb, odom, prod)

    atualizar_progresso_streamlit(100, "Processamento concluido. Excel final gerado")
    registrar_log_streamlit("Arquivo Excel final gerado com sucesso")

    with open(arq_saida, "rb") as f:
        excel_bytes = f.read()

    return excel_bytes, indicadores, resultado_final


def app_streamlit():
    st.set_page_config(
        page_title="ODOMETRO_MATCH - Maxtrack + Cliente",
        page_icon="🚚",
        layout="wide",
    )

    st.title("🚚 Gerador de ODOMETRO_MATCH")
    st.caption("Maxtrack como fonte principal. Cliente/Producao como complemento. Exportacao final em Excel.")

    with st.sidebar:
        st.header("Regras mantidas")
        st.markdown(
            """
- **ODOMETRO_MATCH** e a unica coluna oficial de odometro.
- **Maxtrack** e a fonte principal quando houver match temporal.
- **Cliente/Producao** complementa quando faltar Maxtrack.
- Trava de salto: **1500 km**.
- Cliente acima de **700 km no dia** ativa tratamento de possivel multiplas viagens.
- AM/PM e usado apenas internamente, sem gerar coluna no Excel.
            """
        )
        st.divider()
        st.header("Abas esperadas")
        st.write(f"Combustivel: `{ABA_COMBUSTIVEL}`")
        st.write(f"Maxtrack: `{ABA_MAXTRACK}`")
        st.write(f"Ativo: `{ABA_ATIVO}`")
        st.write(f"Cliente/Producao: `{ABA_PRODUCAO}`")

    st.subheader("1. Selecione as 4 bases")
    col1, col2 = st.columns(2)

    with col1:
        st.info("1 de 4 - Selecione a **Base de Combustivel**")
        arq_combustivel_up = st.file_uploader(
            "Arquivo esperado: Base de dados - Combustivel.xlsx",
            type=["xlsx", "xls"],
            key="combustivel",
        )

        st.info("2 de 4 - Selecione a **Base Maxtrack / KM Rodado**")
        arq_maxtrack_up = st.file_uploader(
            "Arquivo esperado: Base de dados - Km Rodado Maxtrack.xlsx",
            type=["xlsx", "xls"],
            key="maxtrack",
        )

    with col2:
        st.info("3 de 4 - Selecione a **Base de Ativos / Veiculos**")
        arq_ativo_up = st.file_uploader(
            "Arquivo esperado: Base de dados - Ativo de Veiculos.xlsx",
            type=["xlsx", "xls"],
            key="ativo",
        )

        st.info("4 de 4 - Selecione a **Base Cliente / Producao Oficial**")
        arq_producao_up = st.file_uploader(
            "Arquivo esperado: Producao Oficial - Relatorio Cliente.xlsx",
            type=["xlsx", "xls"],
            key="producao",
        )

    nome_saida = st.text_input(
        "Nome do arquivo final para download",
        value="Resumo_Abastecimento_Odometro_Streamlit.xlsx",
    )

    faltantes = validar_uploads(arq_combustivel_up, arq_maxtrack_up, arq_ativo_up, arq_producao_up)
    if faltantes:
        st.warning("Arquivos pendentes:\n\n" + "\n".join([f"- {x}" for x in faltantes]))
    else:
        st.success("As 4 bases foram selecionadas. Clique em processar para gerar o Excel final.")

    st.subheader("2. Processamento")
    processar = st.button("▶️ Processar e gerar Excel", type="primary", disabled=bool(faltantes))

    if processar:
        try:
            with st.spinner("Processando bases e aplicando regras de ODOMETRO_MATCH..."):
                excel_bytes, indicadores, resultado_final = processar_streamlit(
                    arq_combustivel_up,
                    arq_maxtrack_up,
                    arq_ativo_up,
                    arq_producao_up,
                    nome_saida,
                )

            st.success("Processamento concluido com sucesso.")

            st.subheader("3. Indicadores finais")
            st.dataframe(indicadores, use_container_width=True, hide_index=True)

            c1, c2, c3, c4 = st.columns(4)
            with c1:
                st.metric("Registros finais", f"{len(resultado_final):,}".replace(",", "."))
            with c2:
                st.metric("ODOMETRO_MATCH preenchido", f"{int(resultado_final['ODOMETRO_MATCH'].notna().sum()):,}".replace(",", "."))
            with c3:
                st.metric("Saltos > 1500", int((resultado_final["KM_ENTRE_ABAST"] > MAX_SALTO_ABAST).sum()))
            with c4:
                st.metric("Abaixo Cliente ajustado", int(((resultado_final["DIFERENCA_ODOM_X_CLIENTE"] < -1e-9) & resultado_final["DIFERENCA_ODOM_X_CLIENTE"].notna()).sum()))

            st.download_button(
                label="⬇️ Baixar Excel final",
                data=excel_bytes,
                file_name=nome_saida if nome_saida.lower().endswith(".xlsx") else f"{nome_saida}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )

        except Exception as e:
            st.error("Erro durante o processamento. Verifique se os arquivos, abas e colunas estao corretos.")
            st.exception(e)
            st.code(traceback.format_exc(), language="text")


if __name__ == "__main__":
    app_streamlit()
