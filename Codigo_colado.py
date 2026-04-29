import sys
import os
import re
from pathlib import Path
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# =========================================================
# CONFIGURAÇÕES
# =========================================================
TEMPO_MINIMO = 15
TEMPO_MAXIMO = 55
TEMPO_MAXIMO_FALLBACK_SEG = 24 * 3600
ANO_MAXIMO_ACEITO = 2035

COLUNAS_OBRIGATORIAS = [
    "Data",
    "Placa",
    "Tela",
    "Nome da operação",
    "Detalhes"
]

PALAVRAS_CABECALHO = {
    "Data",
    "Placa",
    "Tela",
    "Nome da operação",
    "Detalhes"
}


# =========================================================
# LOG
# =========================================================
def log(msg):
    agora = datetime.now().strftime("%H:%M:%S")
    print(f"[{agora}] {msg}")


def linha():
    print("-" * 110)


# =========================================================
# ENTRADA
# =========================================================
def selecionar_arquivo():
    if len(sys.argv) > 1:
        caminho = sys.argv[1].strip().strip('"')

        if os.path.exists(caminho):
            return caminho

        candidato = os.path.join(os.getcwd(), caminho)
        if os.path.exists(candidato):
            return candidato

        log(f"Arquivo informado não encontrado: {caminho}")
        return None

    arquivos = sorted([
        f for f in os.listdir()
        if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~$")
    ])

    if not arquivos:
        log("Nenhum arquivo Excel encontrado na pasta atual.")
        return None

    linha()
    print("ARQUIVOS DISPONÍVEIS:\n")
    for i, arq in enumerate(arquivos, start=1):
        print(f"{i:02d} - {arq}")
    linha()

    escolha = input("Digite o número do arquivo desejado: ").strip()

    try:
        idx = int(escolha) - 1
        if idx < 0 or idx >= len(arquivos):
            raise ValueError
        return arquivos[idx]
    except Exception:
        log("Escolha inválida.")
        return None


# =========================================================
# UTILITÁRIOS
# =========================================================
def normalizar_texto(valor):
    if pd.isna(valor):
        return ""
    valor = str(valor).strip().upper()
    valor = re.sub(r"\s+", " ", valor)
    return valor


def limpar_coluna_nome(col):
    if pd.isna(col):
        return ""
    col = str(col).strip()
    col = re.sub(r"\s+", " ", col)
    return col


def formatar_data_br(valor):
    if pd.isna(valor):
        return None
    try:
        return pd.to_datetime(valor).strftime("%d/%m/%Y %H:%M:%S")
    except Exception:
        return None


def formatar_duracao_hms(segundos):
    if pd.isna(segundos):
        return None
    try:
        segundos = int(round(float(segundos)))
        if segundos < 0:
            return None
        return str(timedelta(seconds=segundos))
    except Exception:
        return None


def data_valida(dt):
    if pd.isna(dt):
        return False
    if dt.year > ANO_MAXIMO_ACEITO:
        return False
    return True


def tempo_macro_valido(segundos):
    if pd.isna(segundos):
        return False
    try:
        segundos = float(segundos)
    except Exception:
        return False

    if segundos <= 0:
        return False
    if segundos > TEMPO_MAXIMO_FALLBACK_SEG:
        return False

    return True


def extrair_campo_detalhes(texto, nome_campo):
    if pd.isna(texto):
        return None

    texto = str(texto)
    padrao = rf"{re.escape(nome_campo)}\s*::\s*(.*?)(?:\r?\n|$)"
    m = re.search(padrao, texto, flags=re.IGNORECASE)

    if m:
        return m.group(1).strip()
    return None


def classificar_status_por_minutos(minutos):
    if minutos is None or pd.isna(minutos):
        return "IMPROCEDENTE"
    return "OK" if TEMPO_MINIMO <= minutos <= TEMPO_MAXIMO else "IMPROCEDENTE"


def classificar_faixa_tempo(minutos):
    if minutos is None or pd.isna(minutos):
        return "SEM_TEMPO_VALIDO"
    if minutos < TEMPO_MINIMO:
        return f"MENOR_QUE_{TEMPO_MINIMO}"
    if TEMPO_MINIMO <= minutos <= TEMPO_MAXIMO:
        return f"ENTRE_{TEMPO_MINIMO}_E_{TEMPO_MAXIMO}"
    return f"MAIOR_QUE_{TEMPO_MAXIMO}"


def classificar_confiabilidade(origem):
    if origem == "INICIO_FIM":
        return "ALTA"
    if origem == "TEMPO_MACRO":
        return "MEDIA"
    return "BAIXA"


def classificar_motivo(minutos, origem):
    if minutos is None or pd.isna(minutos):
        return "SEM_TEMPO_VALIDO"
    if minutos < TEMPO_MINIMO:
        return "ABAIXO_DO_MINIMO"
    if minutos > TEMPO_MAXIMO:
        return "ACIMA_DO_MAXIMO"
    if origem == "INICIO_FIM":
        return "DENTRO_DA_FAIXA_INICIO_FIM"
    if origem == "TEMPO_MACRO":
        return "DENTRO_DA_FAIXA_TEMPO_MACRO"
    return "DENTRO_DA_FAIXA"


def obter_nome_saida_livre(saida_base):
    saida_base = Path(saida_base)

    if not saida_base.exists():
        return saida_base

    i = 1
    while True:
        novo = saida_base.with_name(f"{saida_base.stem}_{i}{saida_base.suffix}")
        if not novo.exists():
            return novo
        i += 1


# =========================================================
# LEITURA
# =========================================================
def encontrar_linha_cabecalho(arquivo_excel, aba=0, limite_linhas=25):
    previa = pd.read_excel(arquivo_excel, sheet_name=aba, header=None, nrows=limite_linhas)

    melhor_linha = None
    melhor_score = -1

    for i in range(len(previa)):
        valores = set(
            limpar_coluna_nome(x) for x in previa.iloc[i].tolist() if pd.notna(x)
        )
        score = len(valores.intersection(PALAVRAS_CABECALHO))

        if score > melhor_score:
            melhor_score = score
            melhor_linha = i

    if melhor_linha is None:
        raise ValueError("Não foi possível detectar a linha de cabeçalho.")

    log(f"Cabeçalho detectado na linha Excel: {melhor_linha + 1}")
    return melhor_linha


def carregar_dados(arquivo_excel):
    log(f"Lendo arquivo: {arquivo_excel}")

    xls = pd.ExcelFile(arquivo_excel)
    abas = xls.sheet_names

    if not abas:
        raise ValueError("Nenhuma aba encontrada no arquivo.")

    aba = abas[0]
    log(f"Aba selecionada: {aba}")

    linha_cab = encontrar_linha_cabecalho(arquivo_excel, aba=aba)

    df = pd.read_excel(arquivo_excel, sheet_name=aba, header=linha_cab)
    df = df.dropna(how="all").copy()
    df.columns = [limpar_coluna_nome(c) for c in df.columns]

    faltando = [c for c in COLUNAS_OBRIGATORIAS if c not in df.columns]
    if faltando:
        raise ValueError(f"Colunas obrigatórias não encontradas: {faltando}")

    log(f"Total de linhas lidas: {len(df)}")
    log(f"Total de colunas lidas: {len(df.columns)}")

    df["Data_dt"] = pd.to_datetime(df["Data"], dayfirst=True, errors="coerce")
    df["Data_evento_original"] = df["Data_dt"].apply(formatar_data_br)

    df["Tela_norm"] = df["Tela"].apply(normalizar_texto)
    df["Operacao_norm"] = df["Nome da operação"].apply(normalizar_texto)

    if "Tempo de Permanência (segundos)" in df.columns:
        df["Tempo_macro_seg"] = pd.to_numeric(df["Tempo de Permanência (segundos)"], errors="coerce")
        df.loc[
            (df["Tempo_macro_seg"] <= 0) |
            (df["Tempo_macro_seg"] > TEMPO_MAXIMO_FALLBACK_SEG),
            "Tempo_macro_seg"
        ] = pd.NA
    else:
        df["Tempo_macro_seg"] = pd.NA

    df["Tempo_macro_hms"] = df["Tempo_macro_seg"].apply(formatar_duracao_hms)

    df["UP"] = df["Detalhes"].apply(lambda x: extrair_campo_detalhes(x, "Unidade de producao"))
    df["Material"] = df["Detalhes"].apply(lambda x: extrair_campo_detalhes(x, "Material a ser carregado"))
    df["Veiculo_carregado"] = df["Detalhes"].apply(lambda x: extrair_campo_detalhes(x, "Veiculo a ser carregado"))

    return df


# =========================================================
# EVENTOS
# =========================================================
def identificar_eventos_carregamento(df):
    inicio_mask = (
        df["Tela_norm"].str.contains("INICIO DE CARREGAMENTO", na=False) |
        df["Operacao_norm"].eq("CARREGANDO")
    )

    fim_mask = (
        df["Tela_norm"].str.contains("FIM DE CARREGAMENTO", na=False) |
        df["Operacao_norm"].str.contains("FIM CARREGAMENTO", na=False)
    )

    macro_mask = (
        df["Tela_norm"].str.contains("CARREGAMENTO", na=False) |
        df["Operacao_norm"].str.contains("CARREGANDO", na=False) |
        df["Operacao_norm"].str.contains("FIM CARREGAMENTO", na=False)
    )

    eventos = df[inicio_mask | fim_mask | macro_mask].copy()
    eventos = eventos.sort_values(["Placa", "Data_dt"]).reset_index(drop=False)

    log(f"Eventos relacionados a carregamento encontrados: {len(eventos)}")
    return eventos


def classificar_tipo_evento(linha_evento):
    tela = linha_evento.get("Tela_norm", "")
    oper = linha_evento.get("Operacao_norm", "")

    if "INICIO DE CARREGAMENTO" in tela:
        return "INICIO"

    if "FIM DE CARREGAMENTO" in tela:
        return "FIM"

    if "FIM CARREGAMENTO" in oper:
        return "FIM"

    if oper == "CARREGANDO":
        return "INICIO"

    if "CARREGAMENTO" in tela or "CARREGANDO" in oper:
        return "MACRO"

    return None


# =========================================================
# RESULTADOS
# =========================================================
def montar_ciclos_carregamento(eventos):
    registros = []
    alertas = []
    usados = set()

    for placa, grupo in eventos.groupby("Placa", dropna=False):
        grupo = grupo.sort_values("Data_dt").reset_index(drop=True)
        inicio_aberto = None

        for _, linha_evento in grupo.iterrows():
            idx_original = linha_evento["index"]
            tipo = classificar_tipo_evento(linha_evento)

            if tipo == "INICIO":
                if inicio_aberto is not None:
                    alertas.append({
                        "Placa": placa,
                        "Tipo_alerta": "INICIO_DUPLO",
                        "Data_alerta": linha_evento.get("Data_evento_original"),
                        "Observacao": "Novo início encontrado antes do fechamento do anterior."
                    })
                inicio_aberto = linha_evento

            elif tipo == "FIM":
                if inicio_aberto is None:
                    alertas.append({
                        "Placa": placa,
                        "Tipo_alerta": "FIM_SEM_INICIO",
                        "Data_alerta": linha_evento.get("Data_evento_original"),
                        "Observacao": "Fim encontrado sem início correspondente."
                    })
                    continue

                dt_inicio = inicio_aberto.get("Data_dt")
                dt_fim = linha_evento.get("Data_dt")

                if not data_valida(dt_inicio) or not data_valida(dt_fim):
                    alertas.append({
                        "Placa": placa,
                        "Tipo_alerta": "DATA_INVALIDA",
                        "Data_alerta": linha_evento.get("Data_evento_original"),
                        "Observacao": "Data inválida ou fora da faixa aceitável."
                    })
                    inicio_aberto = None
                    continue

                if dt_fim < dt_inicio:
                    alertas.append({
                        "Placa": placa,
                        "Tipo_alerta": "FIM_ANTES_DO_INICIO",
                        "Data_alerta": linha_evento.get("Data_evento_original"),
                        "Observacao": "Fim anterior ao início."
                    })
                    inicio_aberto = None
                    continue

                tempo_segundos = int((dt_fim - dt_inicio).total_seconds())

                if tempo_segundos <= 0 or tempo_segundos > TEMPO_MAXIMO_FALLBACK_SEG:
                    alertas.append({
                        "Placa": placa,
                        "Tipo_alerta": "DURACAO_INVALIDA",
                        "Data_alerta": linha_evento.get("Data_evento_original"),
                        "Observacao": "Duração inválida no par início/fim."
                    })
                    inicio_aberto = None
                    continue

                tempo_minutos_decimal = round(tempo_segundos / 60.0, 2)
                origem = "INICIO_FIM"

                registros.append({
                    "Placa_equipamento": placa,
                    "UP": inicio_aberto.get("UP"),
                    "Material": inicio_aberto.get("Material"),
                    "Veiculo_carregado": inicio_aberto.get("Veiculo_carregado"),
                    "Inicio_carregamento": formatar_data_br(dt_inicio),
                    "Fim_carregamento": formatar_data_br(dt_fim),
                    "Tempo_carregamento": formatar_duracao_hms(tempo_segundos),
                    "Tempo_segundos": tempo_segundos,
                    "Tempo_minutos_decimal": tempo_minutos_decimal,
                    "Faixa_tempo": classificar_faixa_tempo(tempo_minutos_decimal),
                    "Status": classificar_status_por_minutos(tempo_minutos_decimal),
                    "Confiabilidade": classificar_confiabilidade(origem),
                    "Origem_tempo": origem,
                    "Motivo_classificacao": classificar_motivo(tempo_minutos_decimal, origem),
                    "Data_inicio_original": inicio_aberto.get("Data_evento_original"),
                    "Data_fim_original": linha_evento.get("Data_evento_original"),
                    "Tela_inicio": inicio_aberto.get("Tela"),
                    "Operacao_inicio": inicio_aberto.get("Nome da operação"),
                    "Tela_fim": linha_evento.get("Tela"),
                    "Operacao_fim": linha_evento.get("Nome da operação"),
                    "Detalhes_bruto": inicio_aberto.get("Detalhes"),
                })

                usados.add(inicio_aberto["index"])
                usados.add(idx_original)
                inicio_aberto = None

        if inicio_aberto is not None:
            alertas.append({
                "Placa": placa,
                "Tipo_alerta": "INICIO_SEM_FIM",
                "Data_alerta": inicio_aberto.get("Data_evento_original"),
                "Observacao": "Início ficou aberto sem evento final correspondente."
            })

    # fallback: tempo da macro
    for _, linha_evento in eventos.iterrows():
        idx = linha_evento["index"]
        tipo = classificar_tipo_evento(linha_evento)

        if idx in usados:
            continue

        if tipo in ("MACRO", "INICIO", "FIM"):
            tempo_macro_seg = linha_evento.get("Tempo_macro_seg")

            if tempo_macro_valido(tempo_macro_seg):
                tempo_macro_seg = int(round(float(tempo_macro_seg)))
                dt_inicio = linha_evento.get("Data_dt")

                if not data_valida(dt_inicio):
                    alertas.append({
                        "Placa": linha_evento.get("Placa"),
                        "Tipo_alerta": "DATA_INVALIDA_MACRO",
                        "Data_alerta": linha_evento.get("Data_evento_original"),
                        "Observacao": "Data inválida para uso do tempo da macro."
                    })
                    continue

                dt_fim = dt_inicio + timedelta(seconds=tempo_macro_seg)

                if not data_valida(dt_fim):
                    alertas.append({
                        "Placa": linha_evento.get("Placa"),
                        "Tipo_alerta": "FIM_MACRO_INVALIDO",
                        "Data_alerta": linha_evento.get("Data_evento_original"),
                        "Observacao": "Fim calculado pela macro ficou fora da faixa aceitável."
                    })
                    continue

                tempo_minutos_decimal = round(tempo_macro_seg / 60.0, 2)
                origem = "TEMPO_MACRO"

                registros.append({
                    "Placa_equipamento": linha_evento.get("Placa"),
                    "UP": linha_evento.get("UP"),
                    "Material": linha_evento.get("Material"),
                    "Veiculo_carregado": linha_evento.get("Veiculo_carregado"),
                    "Inicio_carregamento": formatar_data_br(dt_inicio),
                    "Fim_carregamento": formatar_data_br(dt_fim),
                    "Tempo_carregamento": formatar_duracao_hms(tempo_macro_seg),
                    "Tempo_segundos": tempo_macro_seg,
                    "Tempo_minutos_decimal": tempo_minutos_decimal,
                    "Faixa_tempo": classificar_faixa_tempo(tempo_minutos_decimal),
                    "Status": classificar_status_por_minutos(tempo_minutos_decimal),
                    "Confiabilidade": classificar_confiabilidade(origem),
                    "Origem_tempo": origem,
                    "Motivo_classificacao": classificar_motivo(tempo_minutos_decimal, origem),
                    "Data_inicio_original": linha_evento.get("Data_evento_original"),
                    "Data_fim_original": None,
                    "Tela_inicio": linha_evento.get("Tela"),
                    "Operacao_inicio": linha_evento.get("Nome da operação"),
                    "Tela_fim": None,
                    "Operacao_fim": None,
                    "Detalhes_bruto": linha_evento.get("Detalhes"),
                })

                usados.add(idx)

    df_resultado = pd.DataFrame(registros)
    df_alertas = pd.DataFrame(alertas)

    if not df_resultado.empty:
        df_resultado = df_resultado.sort_values(
            ["Placa_equipamento", "Inicio_carregamento", "Fim_carregamento"],
            na_position="last"
        ).reset_index(drop=True)

    return df_resultado, df_alertas


# =========================================================
# RESUMOS
# =========================================================
def gerar_resumos(df_resultado):
    if df_resultado.empty:
        resumo_geral = pd.DataFrame([{
            "Total_eventos": 0,
            "Eventos_OK": 0,
            "Eventos_IMPROCEDENTES": 0,
            "Tempo_medio_min": 0,
            "Tempo_mediano_min": 0,
            "Tempo_minimo_min": 0,
            "Tempo_maximo_min": 0
        }])
        return resumo_geral, pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    resumo_geral = pd.DataFrame([{
        "Total_eventos": len(df_resultado),
        "Eventos_OK": int((df_resultado["Status"] == "OK").sum()),
        "Eventos_IMPROCEDENTES": int((df_resultado["Status"] == "IMPROCEDENTE").sum()),
        "Tempo_medio_min": round(df_resultado["Tempo_minutos_decimal"].mean(), 2),
        "Tempo_mediano_min": round(df_resultado["Tempo_minutos_decimal"].median(), 2),
        "Tempo_minimo_min": round(df_resultado["Tempo_minutos_decimal"].min(), 2),
        "Tempo_maximo_min": round(df_resultado["Tempo_minutos_decimal"].max(), 2),
    }])

    resumo_por_up = (
        df_resultado.groupby("UP", dropna=False)
        .agg(
            Total_eventos=("UP", "size"),
            Eventos_OK=("Status", lambda s: int((s == "OK").sum())),
            Eventos_IMPROCEDENTES=("Status", lambda s: int((s == "IMPROCEDENTE").sum())),
            Tempo_medio_min=("Tempo_minutos_decimal", "mean"),
            Tempo_minimo_min=("Tempo_minutos_decimal", "min"),
            Tempo_maximo_min=("Tempo_minutos_decimal", "max"),
        )
        .reset_index()
    )

    resumo_por_equipamento = (
        df_resultado.groupby("Placa_equipamento", dropna=False)
        .agg(
            Total_eventos=("Placa_equipamento", "size"),
            Eventos_OK=("Status", lambda s: int((s == "OK").sum())),
            Eventos_IMPROCEDENTES=("Status", lambda s: int((s == "IMPROCEDENTE").sum())),
            Tempo_medio_min=("Tempo_minutos_decimal", "mean"),
            Tempo_minimo_min=("Tempo_minutos_decimal", "min"),
            Tempo_maximo_min=("Tempo_minutos_decimal", "max"),
        )
        .reset_index()
    )

    ranking_improcedentes = (
        df_resultado[df_resultado["Status"] == "IMPROCEDENTE"]
        .sort_values(["Tempo_minutos_decimal", "Placa_equipamento"], ascending=[False, True])
        .reset_index(drop=True)
    )

    for tabela in [resumo_por_up, resumo_por_equipamento]:
        for col in ["Tempo_medio_min", "Tempo_minimo_min", "Tempo_maximo_min"]:
            if col in tabela.columns:
                tabela[col] = tabela[col].round(2)

    return resumo_geral, resumo_por_up, resumo_por_equipamento, ranking_improcedentes


# =========================================================
# FORMATAÇÃO EXCEL
# =========================================================
def aplicar_estilo_planilha(ws):
    ws.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    max_col = ws.max_column
    max_row = ws.max_row

    if max_row >= 2 and max_col >= 1:
        ref = f"A1:{get_column_letter(max_col)}{max_row}"
        nome_tabela = f"Tabela_{re.sub(r'[^A-Za-z0-9_]', '_', ws.title)}"
        try:
            tabela = Table(displayName=nome_tabela[:255], ref=ref)
            estilo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            tabela.tableStyleInfo = estilo
            ws.add_table(tabela)
        except Exception:
            pass

    for col_cells in ws.columns:
        comprimento = 0
        col_letter = get_column_letter(col_cells[0].column)

        for cell in col_cells:
            try:
                valor = "" if cell.value is None else str(cell.value)
                if len(valor) > comprimento:
                    comprimento = len(valor)
            except Exception:
                pass

        ws.column_dimensions[col_letter].width = min(comprimento + 2, 50)


def colorir_resultados(ws):
    headers = [cell.value for cell in ws[1]]
    if "Status" not in headers or "Origem_tempo" not in headers:
        return

    idx_status = headers.index("Status") + 1
    idx_origem = headers.index("Origem_tempo") + 1

    fill_ok = PatternFill("solid", fgColor="C6EFCE")
    fill_improcedente = PatternFill("solid", fgColor="FFC7CE")
    fill_macro = PatternFill("solid", fgColor="FFF2CC")

    for row in range(2, ws.max_row + 1):
        status = ws.cell(row=row, column=idx_status).value
        origem = ws.cell(row=row, column=idx_origem).value

        if status == "OK":
            fill = fill_ok
        else:
            fill = fill_improcedente

        if origem == "TEMPO_MACRO":
            fill = fill_macro if status == "OK" else fill_improcedente

        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill


# =========================================================
# SAÍDA
# =========================================================
def salvar_saida(arquivo_entrada, df_base, eventos_brutos, df_resultado, df_alertas,
                 resumo_geral, resumo_por_up, resumo_por_equipamento, ranking_improcedentes):
    caminho = Path(arquivo_entrada)
    saida_base = caminho.with_name(f"{caminho.stem}_V3_Profissional.xlsx")
    saida = obter_nome_saida_livre(saida_base)

    try:
        with pd.ExcelWriter(saida, engine="openpyxl") as writer:
            df_base.to_excel(writer, sheet_name="01_Base_Tratada", index=False)
            eventos_brutos.to_excel(writer, sheet_name="02_Eventos_Carregamento", index=False)
            df_resultado.to_excel(writer, sheet_name="03_Resultados", index=False)
            df_alertas.to_excel(writer, sheet_name="04_Alertas", index=False)
            resumo_geral.to_excel(writer, sheet_name="05_Resumo_Geral", index=False)
            resumo_por_up.to_excel(writer, sheet_name="06_Resumo_UP", index=False)
            resumo_por_equipamento.to_excel(writer, sheet_name="07_Resumo_Equipamento", index=False)
            ranking_improcedentes.to_excel(writer, sheet_name="08_Ranking_Improcedentes", index=False)
    except PermissionError:
        saida = obter_nome_saida_livre(caminho.with_name(f"{caminho.stem}_V3_Profissional_fecharexcel.xlsx"))
        with pd.ExcelWriter(saida, engine="openpyxl") as writer:
            df_base.to_excel(writer, sheet_name="01_Base_Tratada", index=False)
            eventos_brutos.to_excel(writer, sheet_name="02_Eventos_Carregamento", index=False)
            df_resultado.to_excel(writer, sheet_name="03_Resultados", index=False)
            df_alertas.to_excel(writer, sheet_name="04_Alertas", index=False)
            resumo_geral.to_excel(writer, sheet_name="05_Resumo_Geral", index=False)
            resumo_por_up.to_excel(writer, sheet_name="06_Resumo_UP", index=False)
            resumo_por_equipamento.to_excel(writer, sheet_name="07_Resumo_Equipamento", index=False)
            ranking_improcedentes.to_excel(writer, sheet_name="08_Ranking_Improcedentes", index=False)

    wb = load_workbook(saida)

    for ws in wb.worksheets:
        aplicar_estilo_planilha(ws)

    if "03_Resultados" in wb.sheetnames:
        colorir_resultados(wb["03_Resultados"])

    wb.save(saida)
    return str(saida)


# =========================================================
# RESUMO TERMINAL
# =========================================================
def imprimir_resumo(resumo_geral, df_alertas):
    linha()
    print("RESUMO GERAL\n")

    if resumo_geral.empty:
        print("Sem dados.")
        linha()
        return

    r = resumo_geral.iloc[0]
    print(f"Total de eventos         : {r['Total_eventos']}")
    print(f"Eventos OK               : {r['Eventos_OK']}")
    print(f"Eventos improcedentes    : {r['Eventos_IMPROCEDENTES']}")
    print(f"Tempo médio (min)        : {r['Tempo_medio_min']}")
    print(f"Tempo mediano (min)      : {r['Tempo_mediano_min']}")
    print(f"Tempo mínimo (min)       : {r['Tempo_minimo_min']}")
    print(f"Tempo máximo (min)       : {r['Tempo_maximo_min']}")
    print(f"Alertas operacionais     : {0 if df_alertas.empty else len(df_alertas)}")
    linha()


# =========================================================
# MAIN
# =========================================================
def main():
    try:
        linha()
        log("INÍCIO DA EXECUÇÃO")

        arquivo = selecionar_arquivo()
        if not arquivo:
            log("Nenhum arquivo selecionado.")
            return

        log(f"Arquivo escolhido: {arquivo}")

        df_base = carregar_dados(arquivo)
        eventos_brutos = identificar_eventos_carregamento(df_base)
        df_resultado, df_alertas = montar_ciclos_carregamento(eventos_brutos)

        log(f"Resultados gerados: {len(df_resultado)}")
        log(f"Alertas encontrados: {0 if df_alertas.empty else len(df_alertas)}")

        resumo_geral, resumo_por_up, resumo_por_equipamento, ranking_improcedentes = gerar_resumos(df_resultado)

        imprimir_resumo(resumo_geral, df_alertas)

        arquivo_saida = salvar_saida(
            arquivo_entrada=arquivo,
            df_base=df_base,
            eventos_brutos=eventos_brutos,
            df_resultado=df_resultado,
            df_alertas=df_alertas,
            resumo_geral=resumo_geral,
            resumo_por_up=resumo_por_up,
            resumo_por_equipamento=resumo_por_equipamento,
            ranking_improcedentes=ranking_improcedentes
        )

        log("Arquivo final salvo com sucesso:")
        print(arquivo_saida)

        log("FIM DA EXECUÇÃO")

    except Exception as e:
        linha()
        log(f"ERRO: {e}")
        linha()
        raise


if __name__ == "__main__":
    main()